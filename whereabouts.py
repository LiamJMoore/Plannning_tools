#!/usr/bin/env python3
"""
WHEREABOUTS VALUES ANALYZER v1.5 -- SPEN
===============================================================================
Sheets (8):
  1. League Table  2. PM Comparison  3. Value Analysis  4. By Work Type
  5. Daily Breakdown  6. Scheme Analysis  7. All Jobs (Sorted)  8. Raw Data

Value Analysis includes: gang run rates, variance, missing whereabouts
  alert (per-gang + individual jobs), value distribution, finish status,
  top 10 jobs, and EVERY job listed by day with values.

Share: Clipboard (text) | Outlook (HTML email + attachment) | Teams webhook
Requirements: pip install pandas openpyxl
"""

import pandas as pd
import sys, os, re, traceback
from datetime import datetime
from pathlib import Path
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
    HAS_TK = True
except ImportError:
    HAS_TK = False

P = {
    'dk': '1B2A4A', 'mid': '2C3E6B', 'accent': 'F4B942',
    'g': 'C6EFCE', 'a': 'FFF2CC', 'r': 'FFC7CE', 'b': 'D6E4F0',
    'alt': 'F0F3F8', 'wh': 'FFFFFF', 'bdr': 'D5D8DC',
    'gold': 'FFD700', 'silver': 'C0C0C0', 'bronze': 'CD7F32',
    'kpi_g': 'D4F7C2', 'kpi_b': 'D5E1F7', 'kpi_a': 'FCE4D6', 'kpi_p': 'E1D5E7',
    'warn': 'FFF0E0',
}
GBP = chr(163)

PLANNED_REPORT_COL_MAP = {
    'Gang Ref': 'Gang',
    'Planned Value': 'Job Value',
    'Measured Value': 'Whereabouts Value',
    'Date From': 'Scheduled From',
    'Job Status': 'Finish Status',
}
PLANNED_REPORT_KEEP = [
    'Week Commencing', 'Remaining Planned Value', 'Previous Measures Logged',
    'Measures (Inc Pending)', 'Measures In Period', 'Planned Value2',
    'Latest Planned Date', 'Unscheduled?', 'Started?', 'Measures Logged?',
    'Site Cleared?', 'Ready To Invoice', 'Aborted?', 'Missing Measures?',
    'Pending Measure', 'Pending Measure Value', 'Cost Variation',
    'Cost Variation Value', 'Outstanding Task', 'On Hold?',
    'Measures Logged Date', 'Supervisor', 'Project Manager', 'Planner',
    'Duration', 'RBD', 'ARBD', 'Sub Status', 'Date To',
    'Total Daily Forecast', 'Estimated Daily Forecast Value',
]
PLANNED_REPORT_DETECT_COLS = {'Gang Ref', 'Planned Value', 'Measured Value'}


# ===================================================================
# ENGINE
# ===================================================================

class Engine:
    def __init__(self):
        self.raw = self.data = None
        self.files_loaded = []
        self.gang_stats = {}; self.pm_stats = {}; self.wt_stats = {}
        self.daily_stats = {}; self.scheme_stats = {}; self.summary = {}
        self.top_jobs = None; self.value_analysis = {}; self.duplicates = []; self.duplicate_inflation = 0; self.duplicate_true_value = 0
        self.finish_stats = {}; self.value_bands = {}
        self.missing_wv = None; self.missing_wv_by_gang = {}; self.missing_wv_by_pm = {}
        self._day_order = []
        self.aborted_raw = None; self.aborted_analysis = None; self.aborted_by_reason = {}; self.aborted_by_gang = {}
        self.planned_stats = {}; self.is_planned_report = False

    @property
    def sorted_days(self):
        """Days in chronological order."""
        return [d for d in self._day_order if d in self.daily_stats]

    @staticmethod
    def _detect_planned_report(df):
        return PLANNED_REPORT_DETECT_COLS.issubset(set(df.columns))

    @staticmethod
    def _remap_planned_report(df):
        keep_extra = [c for c in PLANNED_REPORT_KEEP if c in df.columns]
        df = df.rename(columns=PLANNED_REPORT_COL_MAP)
        return df, keep_extra

    def load_planned_report(self, paths):
        frames = []
        for p in paths:
            xls = pd.ExcelFile(p)
            sheet = None
            for s in xls.sheet_names:
                if 'planned report' in s.lower():
                    sheet = s; break
            if sheet is None:
                df = pd.read_excel(p)
                if not self._detect_planned_report(df):
                    raise ValueError(f"No 'Planned Report' tab found in {os.path.basename(p)} and columns don't match planned format")
            else:
                df = pd.read_excel(p, sheet_name=sheet)
            if df.empty: continue
            if not self._detect_planned_report(df):
                raise ValueError(f"Missing required columns (Gang Ref, Planned Value, Measured Value) in {os.path.basename(p)}")
            df, extra_cols = self._remap_planned_report(df)
            df['_file'] = os.path.basename(p)
            df['_extra_cols'] = ','.join(extra_cols)
            frames.append(df)
            self.files_loaded.append(os.path.basename(p))
        if not frames: raise ValueError("No valid planned report data found")
        self.raw = pd.concat(frames, ignore_index=True).copy()
        self.is_planned_report = True
        return self.raw

    def _planned_analysis(self):
        if not self.is_planned_report or self.data is None: return
        df = self.data
        # Use only first occurrence of each Job ID to avoid double-counting
        if 'Job ID' in df.columns:
            df_unique = df.drop_duplicates(subset='Job ID', keep='first').copy()
        else:
            df_unique = df.copy()
        for col in ['Remaining Planned Value', 'Measures In Period', 'Previous Measures Logged',
                     'Measures (Inc Pending)', 'Pending Measure Value', 'Cost Variation Value',
                     'Total Daily Forecast', 'Estimated Daily Forecast Value']:
            if col in df_unique.columns:
                df_unique[col] = pd.to_numeric(df_unique[col], errors='coerce').fillna(0)
        ps = {}
        total_planned = round(float(df_unique['_orig_jv'].sum()), 2) if '_orig_jv' in df_unique.columns else round(float(df_unique['Job Value'].sum()), 2)
        total_measured = round(float(df_unique['_orig_wv'].sum()), 2) if '_orig_wv' in df_unique.columns else round(float(df_unique['Whereabouts Value'].sum()), 2)
        total_remaining = round(float(df_unique['Remaining Planned Value'].sum()), 2) if 'Remaining Planned Value' in df_unique.columns else 0
        total_measures_period = round(float(df_unique['Measures In Period'].sum()), 2) if 'Measures In Period' in df_unique.columns else 0
        total_prev_measures = round(float(df_unique['Previous Measures Logged'].sum()), 2) if 'Previous Measures Logged' in df_unique.columns else 0
        total_cv_value = round(float(df_unique['Cost Variation Value'].sum()), 2) if 'Cost Variation Value' in df_unique.columns else 0
        total_pending_val = round(float(df_unique['Pending Measure Value'].sum()), 2) if 'Pending Measure Value' in df_unique.columns else 0
        n_aborted = int((df_unique['Aborted?'] == 'Yes').sum()) if 'Aborted?' in df_unique.columns else 0
        n_site_cleared = int((df_unique['Site Cleared?'] == 'Yes').sum()) if 'Site Cleared?' in df_unique.columns else 0
        n_started = int((df_unique['Started?'] == 'Yes').sum()) if 'Started?' in df_unique.columns else 0
        n_measures_logged = int((df_unique['Measures Logged?'] == 'Yes').sum()) if 'Measures Logged?' in df_unique.columns else 0
        n_unscheduled = int((df_unique['Unscheduled?'] == 'Yes').sum()) if 'Unscheduled?' in df_unique.columns else 0
        n_missing_measures = int((df_unique['Missing Measures?'] == 'Yes').sum()) if 'Missing Measures?' in df_unique.columns else 0
        n_cv = int((df_unique['Cost Variation'] == 'Yes').sum()) if 'Cost Variation' in df_unique.columns else 0
        n_on_hold = int((df_unique['On Hold?'] == 'Yes').sum()) if 'On Hold?' in df_unique.columns else 0
        n_ready_invoice = int((df_unique['Ready To Invoice'] == 'Yes').sum()) if 'Ready To Invoice' in df_unique.columns else 0
        n_unique_jobs = len(df_unique)
        measure_rate = round(total_measured / total_planned * 100, 2) if total_planned > 0 else 0
        ps.update({
            'total_planned': total_planned, 'total_measured': total_measured,
            'total_remaining': total_remaining, 'total_measures_period': total_measures_period,
            'total_prev_measures': total_prev_measures, 'total_cv_value': total_cv_value,
            'total_pending_val': total_pending_val,
            'n_aborted': n_aborted, 'n_site_cleared': n_site_cleared,
            'n_started': n_started, 'n_measures_logged': n_measures_logged,
            'n_unscheduled': n_unscheduled, 'n_missing_measures': n_missing_measures,
            'n_cv': n_cv, 'n_on_hold': n_on_hold, 'n_ready_invoice': n_ready_invoice,
            'measure_rate': measure_rate, 'n_unique_jobs': n_unique_jobs,
        })
        # Per-gang planned stats (also using unique jobs only)
        gang_planned = {}
        for gang in df_unique['Gang'].unique():
            gd = df_unique[df_unique['Gang'] == gang]
            gp = {
                'planned': round(float(gd['_orig_jv'].sum()), 2) if '_orig_jv' in gd.columns else round(float(gd['Job Value'].sum()), 2),
                'measured': round(float(gd['_orig_wv'].sum()), 2) if '_orig_wv' in gd.columns else round(float(gd['Whereabouts Value'].sum()), 2),
                'remaining': round(float(gd['Remaining Planned Value'].sum()), 2) if 'Remaining Planned Value' in gd.columns else 0,
                'measures_period': round(float(gd['Measures In Period'].sum()), 2) if 'Measures In Period' in gd.columns else 0,
                'n_aborted': int((gd['Aborted?'] == 'Yes').sum()) if 'Aborted?' in gd.columns else 0,
                'n_site_cleared': int((gd['Site Cleared?'] == 'Yes').sum()) if 'Site Cleared?' in gd.columns else 0,
                'n_started': int((gd['Started?'] == 'Yes').sum()) if 'Started?' in gd.columns else 0,
                'n_missing_measures': int((gd['Missing Measures?'] == 'Yes').sum()) if 'Missing Measures?' in gd.columns else 0,
                'n_cv': int((gd['Cost Variation'] == 'Yes').sum()) if 'Cost Variation' in gd.columns else 0,
                'cv_value': round(float(gd['Cost Variation Value'].sum()), 2) if 'Cost Variation Value' in gd.columns else 0,
            }
            gp['measure_rate'] = round(gp['measured'] / gp['planned'] * 100, 2) if gp['planned'] > 0 else 0
            gang_planned[gang] = gp
        ps['gang_planned'] = gang_planned
        # Job status breakdown (unique jobs only)
        if 'Finish Status' in df_unique.columns:
            ps['status_breakdown'] = {}
            for status in df_unique['Finish Status'].unique():
                sub = df_unique[df_unique['Finish Status'] == status]
                ps['status_breakdown'][status] = {
                    'count': len(sub),
                    'planned': round(float(sub['_orig_jv'].sum()), 2) if '_orig_jv' in sub.columns else round(float(sub['Job Value'].sum()), 2),
                    'measured': round(float(sub['_orig_wv'].sum()), 2) if '_orig_wv' in sub.columns else round(float(sub['Whereabouts Value'].sum()), 2),
                    'remaining': round(float(sub['Remaining Planned Value'].sum()), 2) if 'Remaining Planned Value' in sub.columns else 0,
                }
        self.planned_stats = ps

    def load_aborted(self, paths):
        """Load aborted jobs file(s)."""
        frames = []
        for p in paths:
            df = pd.read_excel(p)
            if df.empty: continue
            if 'Job ID' not in df.columns: raise ValueError(f"Missing 'Job ID' in {os.path.basename(p)}")
            frames.append(df)
        if not frames: return None
        self.aborted_raw = pd.concat(frames, ignore_index=True).copy()
        return self.aborted_raw

    def load_files(self, paths):
        frames = []
        for p in paths:
            df = pd.read_excel(p)
            if df.empty: continue
            for c in ['Gang', 'Job Value']:
                if c not in df.columns: raise ValueError(f"Missing '{c}' in {os.path.basename(p)}")
            df['_file'] = os.path.basename(p); frames.append(df)
            self.files_loaded.append(os.path.basename(p))
        if not frames: raise ValueError("No valid data found")
        self.raw = pd.concat(frames, ignore_index=True).copy()
        return self.raw

    def analyse(self, df):
        df = df.copy()
        df['Gang'] = df['Gang'].fillna('Unassigned')
        df['Job Value'] = pd.to_numeric(df['Job Value'], errors='coerce').fillna(0)
        df['Whereabouts Value'] = pd.to_numeric(df.get('Whereabouts Value', 0), errors='coerce').fillna(0)
        for c, d in [('Work Type', 'Unknown'), ('Finish Status', '')]:
            if c not in df.columns: df[c] = d
            df[c] = df[c].fillna(d)

        # Store original values before dedup
        df['_orig_jv'] = df['Job Value'].copy()
        df['_orig_wv'] = df['Whereabouts Value'].copy()

        # Shared jobs: first gang keeps the full value, others get zero
        if 'Job ID' in df.columns:
            is_dup = df.duplicated(subset='Job ID', keep='first')
            df['_shared'] = df['Job ID'].map(df['Job ID'].value_counts()) > 1
            df['_share_count'] = df['Job ID'].map(df['Job ID'].value_counts())
            df.loc[is_dup, 'Job Value'] = 0
            df.loc[is_dup, 'Whereabouts Value'] = 0
        else:
            df['_shared'] = False
            df['_share_count'] = 1
        df['_pm'] = df['Gang'].apply(lambda x: m.group(1).strip() if (m := re.search(r'\(([^)]+)\)', str(x))) else '')
        df['_short'] = df['Gang'].apply(lambda x: re.sub(r'\s*\(.*?\)', '', str(x)).strip())
        if 'Scheduled From' in df.columns:
            df['Scheduled From'] = pd.to_datetime(df['Scheduled From'], errors='coerce')
            df['_day'] = df['Scheduled From'].dt.strftime('%d/%m').fillna('Unknown')
            # Build date sort order from actual dates
            day_dates = df.dropna(subset=['Scheduled From']).groupby('_day')['Scheduled From'].min().sort_values()
            self._day_order = list(day_dates.index)
        else:
            df['_day'] = 'Unknown'
            self._day_order = ['Unknown']
        if 'Client Ref 2' in df.columns: df['_scheme'] = df['Client Ref 2'].fillna('No Scheme')
        elif 'Contract Number' in df.columns: df['_scheme'] = df['Contract Number'].fillna('No Contract')
        else: df['_scheme'] = 'Unknown'
        self.data = df
        self._detect_duplicates(); self._gangs(); self._pms(); self._work_types()
        self._daily(); self._schemes(); self._top_jobs(); self._value_analysis()
        self._finish_status(); self._value_distribution(); self._missing_whereabouts()
        self._summary()
        if self.is_planned_report: self._planned_analysis()

    def _detect_duplicates(self):
        if 'Job ID' not in self.data.columns: return
        counts = self.data['Job ID'].value_counts()
        for job_id, count in counts[counts > 1].items():
            rows = self.data[self.data['Job ID'] == job_id]
            gangs = rows['Gang'].unique().tolist()
            gang_shorts = [re.sub(r'\s*\(.*?\)', '', g).strip() for g in gangs]
            pms = rows['_pm'].unique().tolist()
            orig_val = round(float(rows['_orig_jv'].iloc[0]), 2)
            orig_wv = round(float(rows['_orig_wv'].iloc[0]), 2)
            # First row keeps the value
            owner_gang = gang_shorts[0] if gang_shorts else ''
            self.duplicates.append({
                'job_id': job_id, 'count': int(count),
                'gangs': gangs, 'gang_shorts': gang_shorts,
                'pms': [p for p in pms if p],
                'orig_val': orig_val, 'orig_wv': orig_wv,
                'owner': owner_gang,
                'scheme': rows['_scheme'].iloc[0] if '_scheme' in rows.columns else '',
                'work_type': rows['Work Type'].iloc[0] if 'Work Type' in rows.columns else '',
                'day': rows['_day'].iloc[0] if '_day' in rows.columns else '',
                'client_ref': rows['Client Ref 1'].iloc[0] if 'Client Ref 1' in rows.columns else '',
            })
        self.duplicate_inflation = round(sum(d['orig_val'] * (d['count'] - 1) for d in self.duplicates), 2)
        self.duplicate_true_value = round(sum(d['orig_val'] for d in self.duplicates), 2)

    def _val_stats(self, sub):
        n = len(sub); jv = round(float(sub['Job Value'].sum()), 2)
        wv = round(float(sub['Whereabouts Value'].sum()), 2)
        avg = round(jv / n, 2) if n else 0
        return {'n': n, 'jv': jv, 'wv': wv, 'avg': avg,
                'cap': int((sub['Work Type'] == 'Capital Works').sum()),
                'defects': int((sub['Work Type'] == 'Defects').sum()),
                'max_val': round(float(sub['Job Value'].max()), 2) if n else 0,
                'min_val': round(float(sub['Job Value'].min()), 2) if n else 0}

    def _gangs(self):
        for gang in self.data['Gang'].unique():
            g = self.data[self.data['Gang'] == gang]; s = self._val_stats(g)
            s['pm'] = g['_pm'].iloc[0] if len(g) else ''
            s['short'] = g['_short'].iloc[0] if len(g) else ''
            s['schemes'] = sorted(g['_scheme'].unique().tolist(), key=str)
            s['days_active'] = g['_day'].nunique()
            self.gang_stats[gang] = s

    def _pms(self):
        for pm in self.data['_pm'].unique():
            if not pm: continue
            g = self.data[self.data['_pm'] == pm]; s = self._val_stats(g)
            teams = sorted(g['Gang'].unique().tolist())
            tb = [{'team': tm, 'short': self.gang_stats.get(tm, {}).get('short', tm),
                   'n': self.gang_stats.get(tm, {}).get('n', 0), 'jv': self.gang_stats.get(tm, {}).get('jv', 0),
                   'wv': self.gang_stats.get(tm, {}).get('wv', 0), 'avg': self.gang_stats.get(tm, {}).get('avg', 0)}
                  for tm in teams]
            s.update({'teams': teams, 'num_teams': len(teams), 'breakdown': tb})
            self.pm_stats[pm] = s

    def _work_types(self):
        for gang in self.data['Gang'].unique():
            gd = self.data[self.data['Gang'] == gang]
            for wt in gd['Work Type'].unique():
                self.wt_stats[(gang, wt)] = self._val_stats(gd[gd['Work Type'] == wt])

    def _daily(self):
        for day in self._day_order:
            dd = self.data[self.data['_day'] == day]
            d = {'overall': self._val_stats(dd), 'gangs': {}, 'pms': {}}
            for gang in dd['Gang'].unique(): d['gangs'][gang] = self._val_stats(dd[dd['Gang'] == gang])
            for pm in dd['_pm'].unique():
                if pm: d['pms'][pm] = self._val_stats(dd[dd['_pm'] == pm])
            self.daily_stats[day] = d

    def _schemes(self):
        for sch in self.data['_scheme'].unique():
            sd = self.data[self.data['_scheme'] == sch]; s = self._val_stats(sd)
            s.update({'teams': sd['Gang'].nunique(), 'team_list': sorted(sd['Gang'].unique().tolist()),
                      'pms': sorted(sd['_pm'].unique().tolist())})
            self.scheme_stats[sch] = s

    def _top_jobs(self):
        cols = [c for c in ['Job ID', 'Client Ref 1', 'Gang', '_pm', '_scheme', 'Work Type',
                'Job Value', 'Whereabouts Value', 'Address', '_day'] if c in self.data.columns]
        self.top_jobs = self.data.nlargest(10, 'Job Value')[cols].reset_index(drop=True)

    def _value_analysis(self):
        gs = self.gang_stats
        if not gs: return
        overall_avg = round(sum(s['jv'] for s in gs.values()) / sum(s['n'] for s in gs.values()), 2) if sum(s['n'] for s in gs.values()) else 0
        for gang, s in gs.items():
            da = max(s.get('days_active', 1), 1)
            dr = round(s['jv'] / da, 2); var = round(s['avg'] - overall_avg, 2)
            vp = round((var / overall_avg * 100), 2) if overall_avg else 0
            gap = round(s['jv'] - s['wv'], 2); ratio = round((s['wv'] / s['jv'] * 100), 2) if s['jv'] > 0 else 0
            self.value_analysis[gang] = {
                'short': s['short'], 'pm': s['pm'], 'n': s['n'], 'jv': s['jv'], 'wv': s['wv'], 'avg': s['avg'],
                'days_active': da, 'daily_rate': dr, 'variance': var, 'variance_pct': vp,
                'wv_gap': gap, 'wv_ratio': ratio, 'max_val': s['max_val'], 'min_val': s['min_val']}

    def _finish_status(self):
        fs = self.data['Finish Status'].replace('', 'No Status')
        for status, count in fs.value_counts().items():
            sub = self.data[self.data['Finish Status'] == ('' if status == 'No Status' else status)]
            self.finish_stats[status] = {'count': int(count), 'jv': round(float(sub['Job Value'].sum()), 2),
                'wv': round(float(sub['Whereabouts Value'].sum()), 2), 'pct': round(count / len(self.data) * 100, 2)}

    def _value_distribution(self):
        df = self.data; total_jv = round(float(df['Job Value'].sum()), 2)
        for label, lo, hi in [('Under 50', 0, 50), ('50 to 100', 50, 100), ('100 to 250', 100, 250),
                 ('250 to 500', 250, 500), ('500 to 1,000', 500, 1000), ('1,000 to 5,000', 1000, 5000),
                 ('5,000 to 10,000', 5000, 10000), ('Over 10,000', 10000, float('inf'))]:
            sub = df[df['Job Value'] >= lo] if hi == float('inf') else df[(df['Job Value'] >= lo) & (df['Job Value'] < hi)]
            if len(sub) > 0:
                self.value_bands[label] = {'count': len(sub), 'jv': round(float(sub['Job Value'].sum()), 2),
                    'avg': round(float(sub['Job Value'].mean()), 2), 'pct_count': round(len(sub) / len(df) * 100, 2),
                    'pct_value': round(float(sub['Job Value'].sum()) / total_jv * 100, 2) if total_jv > 0 else 0}

    def _missing_whereabouts(self):
        df = self.data; missing = df[(df['Whereabouts Value'] == 0) & (df['Job Value'] > 0)].copy()
        self.missing_wv = missing
        for gang in missing['Gang'].unique():
            gm = missing[missing['Gang'] == gang]; gs = self.gang_stats.get(gang, {})
            tc = gs.get('n', 0)
            self.missing_wv_by_gang[gang] = {
                'short': gs.get('short', gang), 'pm': gs.get('pm', ''),
                'missing_count': len(gm), 'total_count': tc,
                'missing_pct': round(len(gm) / tc * 100, 2) if tc else 0,
                'missing_jv': round(float(gm['Job Value'].sum()), 2), 'gang_jv': gs.get('jv', 0),
                'missing_jv_pct': round(float(gm['Job Value'].sum()) / gs.get('jv', 1) * 100, 2) if gs.get('jv', 0) > 0 else 0}
        for pm in missing['_pm'].unique():
            if not pm: continue
            pm_m = missing[missing['_pm'] == pm]; ps = self.pm_stats.get(pm, {})
            self.missing_wv_by_pm[pm] = {
                'missing_count': len(pm_m), 'total_count': ps.get('n', 0),
                'missing_jv': round(float(pm_m['Job Value'].sum()), 2), 'pm_jv': ps.get('jv', 0),
                'missing_pct': round(len(pm_m) / ps.get('n', 1) * 100, 2) if ps.get('n', 0) > 0 else 0}

    def _summary(self):
        df = self.data; gs = self.gang_stats
        total_n = len(df)
        total_jv = round(float(df['Job Value'].sum()), 2)
        total_wv = round(float(df['Whereabouts Value'].sum()), 2)
        mt = round(float(self.missing_wv['Job Value'].sum()), 2) if self.missing_wv is not None and len(self.missing_wv) > 0 else 0
        self.summary = {
            'total_jobs': total_n, 'total_jv': total_jv, 'total_wv': total_wv,
            'n_gangs': len(gs), 'n_pms': len(self.pm_stats), 'n_days': len(self.daily_stats),
            'days': self.sorted_days,
            'contract': df['Contract'].iloc[0] if 'Contract' in df.columns else 'Unknown',
            'region': df['Region'].dropna().iloc[0] if 'Region' in df.columns and df['Region'].notna().any() else 'Unknown',
            'avg_jv': round(total_jv / total_n, 2) if total_n else 0,
            'n_schemes': len(self.scheme_stats), 'n_duplicates': len(self.duplicates),
            'missing_wv_count': len(self.missing_wv) if self.missing_wv is not None else 0,
            'missing_wv_value': mt,
            'is_planned_report': self.is_planned_report}

    def analyse_aborted(self):
        """Cross-reference aborted jobs with whereabouts data."""
        if self.aborted_raw is None or self.aborted_raw.empty:
            return
        ab = self.aborted_raw.copy()
        ab['Gang'] = ab['Gang'].fillna('Unassigned')
        ab['Abort Reason'] = ab['Abort Reason'].fillna('Unknown')
        if 'Aborted Date' in ab.columns:
            ab['Aborted Date'] = pd.to_datetime(ab['Aborted Date'], errors='coerce')
        if 'Comments' not in ab.columns:
            ab['Comments'] = ''
        ab['Comments'] = ab['Comments'].fillna('')

        # Cross-reference with whereabouts to get values
        wv_lookup = {}
        if self.data is not None:
            for _, row in self.data.iterrows():
                jid = row.get('Job ID')
                if jid and jid not in wv_lookup:
                    wv_lookup[jid] = {
                        'jv': round(float(row.get('_orig_jv', row.get('Job Value', 0))), 2),
                        'wv': round(float(row.get('_orig_wv', row.get('Whereabouts Value', 0))), 2),
                        'in_whereabouts': True,
                    }

        # Build analysis for each aborted job
        records = []
        for _, row in ab.iterrows():
            jid = row['Job ID']
            wv_data = wv_lookup.get(jid, {'jv': 0, 'wv': 0, 'in_whereabouts': False})
            records.append({
                'job_id': jid,
                'client_ref': row.get('Client Ref 1', ''),
                'scheme': row.get('Client Ref 2', ''),
                'work_type': row.get('Work Type', ''),
                'gang': row['Gang'],
                'gang_short': re.sub(r'\s*\(.*?\)', '', str(row['Gang'])).strip(),
                'supervisor': row.get('Supervisors', ''),
                'abort_reason': row['Abort Reason'],
                'comments': str(row.get('Comments', '')).replace('\n', ' ').strip(),
                'job_status': row.get('Job Status', ''),
                'address': row.get('Address', ''),
                'aborted_date': row.get('Aborted Date', ''),
                'on_hold': row.get('On Hold?', ''),
                'jv': wv_data['jv'],
                'wv': wv_data['wv'],
                'in_whereabouts': wv_data['in_whereabouts'],
            })

        self.aborted_analysis = pd.DataFrame(records)

        # By reason
        self.aborted_by_reason = {}
        for reason in ab['Abort Reason'].unique():
            sub = self.aborted_analysis[self.aborted_analysis['abort_reason'] == reason]
            self.aborted_by_reason[reason] = {
                'count': len(sub),
                'jv': round(float(sub['jv'].sum()), 2),
                'in_wv': int(sub['in_whereabouts'].sum()),
            }

        # By gang
        self.aborted_by_gang = {}
        for gang in ab['Gang'].unique():
            sub = self.aborted_analysis[self.aborted_analysis['gang'] == gang]
            self.aborted_by_gang[gang] = {
                'gang_short': re.sub(r'\s*\(.*?\)', '', str(gang)).strip(),
                'count': len(sub),
                'jv': round(float(sub['jv'].sum()), 2),
                'in_wv': int(sub['in_whereabouts'].sum()),
                'reasons': dict(sub['abort_reason'].value_counts()),
            }

        # Update summary
        self.summary['n_aborted'] = len(self.aborted_analysis)
        self.summary['aborted_jv'] = round(float(self.aborted_analysis['jv'].sum()), 2)
        self.summary['aborted_in_wv'] = int(self.aborted_analysis['in_whereabouts'].sum())

    def share_text(self):
        s = self.summary; gs = self.gang_stats
        L = [f"{s['contract']} {s['region']} - Whereabouts Values Report"]
        if s['n_days'] > 1: L.append(f"Period: {s['days'][0]} to {s['days'][-1]} ({s['n_days']} days)")
        else: L.append(f"Date: {s['days'][0] if s['days'] else 'N/A'}")
        L.append(f"Report Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        L.append("")
        L.append(f"Total Jobs: {s['total_jobs']}  |  Job Value: {GBP}{s['total_jv']:,.2f}  |  Whereabouts: {GBP}{s['total_wv']:,.2f}  |  Average: {GBP}{s['avg_jv']:,.2f}")
        if self.is_planned_report and self.planned_stats:
            ps = self.planned_stats
            L.append(f"  Remaining: {GBP}{ps['total_remaining']:,.2f}  |  Measure Rate: {ps['measure_rate']:.1f}%")
            L.append(f"  Site Cleared: {ps['n_site_cleared']}  |  Started: {ps['n_started']}  |  Missing Measures: {ps['n_missing_measures']}  |  Aborted: {ps['n_aborted']}")
        if s['n_duplicates'] > 0: L.append(f"  Note: {s['n_duplicates']} shared jobs -- value assigned to first gang only (no double-counting)")
        if s['missing_wv_count'] > 0: L.append(f"  ALERT: {s['missing_wv_count']} jobs ({GBP}{s['missing_wv_value']:,.2f}) have ZERO whereabouts value")
        L.append("")
        L.append("GANG LEAGUE TABLE:")
        L.append(f"{'#':<4} {'Gang':<28} {'PM':<16} {'Jobs':>5} {'Job Value':>14} {'Whereabouts':>14} {'Average':>12}")
        L.append("-" * 97)
        for i, (gang, st) in enumerate(sorted(gs.items(), key=lambda x: x[1]['jv'], reverse=True), 1):
            L.append(f"{i:<4} {(st['short'] or gang):<28} {st['pm']:<16} {st['n']:>5} {GBP}{st['jv']:>12,.2f} {GBP}{st['wv']:>12,.2f} {GBP}{st['avg']:>10,.2f}")
        L.append("-" * 97)
        L.append(f"{'':4} {'TOTAL':<28} {'':16} {s['total_jobs']:>5} {GBP}{s['total_jv']:>12,.2f} {GBP}{s['total_wv']:>12,.2f} {GBP}{s['avg_jv']:>10,.2f}")
        L.append("")
        if self.daily_stats and s['n_days'] > 1:
            L.append("DAILY SUMMARY:")
            for day in self.sorted_days:
                ds = self.daily_stats[day]['overall']
                L.append(f"  {day}: {ds['n']} jobs  |  {GBP}{ds['jv']:,.2f} total  |  {GBP}{ds['avg']:,.2f} average")
            L.append("")
        if self.aborted_analysis is not None and len(self.aborted_analysis) > 0:
            n_ab = len(self.aborted_analysis)
            n_ov = int(self.aborted_analysis['in_whereabouts'].sum())
            jv_ov = round(float(self.aborted_analysis[self.aborted_analysis['in_whereabouts']]['jv'].sum()), 2)
            L.append(f"ABORTED: {n_ab} jobs aborted | {n_ov} matched ({GBP}{jv_ov:,.2f} value)")
            for reason, rs in sorted(self.aborted_by_reason.items(), key=lambda x: -x[1]['count']):
                L.append(f"  {reason}: {rs['count']}")
            L.append("")
        return '\n'.join(L)

    def share_html(self):
        s = self.summary; gs = self.gang_stats
        period = f'{s["days"][0]} to {s["days"][-1]}' if s['n_days'] > 1 else (s['days'][0] if s['days'] else '')
        h = ['<html><head><meta charset="utf-8"></head>',
             '<body style="font-family:Arial,Helvetica,sans-serif;font-size:14px;color:#333333;margin:0;padding:16px;background:#FFFFFF;">',
             f'<h2 style="color:#1B2A4A;margin:0 0 4px 0;font-size:18px;">{s["contract"]} {s["region"]} - Whereabouts Values</h2>',
             f'<p style="color:#888888;margin:0 0 14px 0;font-size:12px;">{period} | Generated {datetime.now().strftime("%d/%m/%Y %H:%M")}</p>',
             '<table cellpadding="8" cellspacing="0" style="border-collapse:collapse;margin-bottom:14px;"><tr>']
        for lbl, val in [('Total Jobs', f'{s["total_jobs"]}'), ('Job Value', f'&pound;{s["total_jv"]:,.2f}'),
                         ('Whereabouts', f'&pound;{s["total_wv"]:,.2f}'), ('Average', f'&pound;{s["avg_jv"]:,.2f}')]:
            h.append(f'<td style="padding:8px 20px 8px 0;vertical-align:top;"><span style="font-size:11px;color:#888888;">{lbl}</span><br><span style="font-size:18px;font-weight:bold;color:#1B2A4A;">{val}</span></td>')
        h.append('</tr></table>')
        if s['missing_wv_count'] > 0:
            mp = round(s['missing_wv_value'] / s['total_jv'] * 100) if s['total_jv'] > 0 else 0
            h.append(f'<div style="background:#FFF3E0;border-left:4px solid #E67E22;padding:10px 14px;margin:0 0 14px 0;font-size:13px;"><strong style="color:#E67E22;">Missing Whereabouts:</strong> {s["missing_wv_count"]} jobs with &pound;{s["missing_wv_value"]:,.2f} job value ({mp}% of total) have zero whereabouts.')
            if self.missing_wv_by_pm:
                parts = [f'{pm}: {mp["missing_count"]} jobs (&pound;{mp["missing_jv"]:,.2f})' for pm, mp in sorted(self.missing_wv_by_pm.items(), key=lambda x: -x[1]['missing_jv'])]
                h.append(f'<br><span style="color:#888888;font-size:12px;">{" | ".join(parts)}</span>')
            h.append('</div>')
        if s['n_duplicates'] > 0: h.append(f'<p style="color:#CC6600;font-size:12px;margin:0 0 10px 0;">Shared Jobs: {s["n_duplicates"]} jobs on multiple gangs - value assigned to first gang only (no double-counting)</p>')
        bc = 'border:1px solid #E0E0E0;'; bch = 'border:1px solid #1B2A4A;'
        h.append('<table cellpadding="7" cellspacing="0" style="border-collapse:collapse;width:100%;font-size:13px;">')
        h.append(f'<tr style="background:#1B2A4A;color:#FFFFFF;"><th style="text-align:center;width:30px;{bch}">#</th><th style="text-align:left;{bch}">Gang</th><th style="text-align:left;{bch}">PM</th><th style="text-align:center;{bch}">Jobs</th><th style="text-align:right;{bch}">Job Value</th><th style="text-align:right;{bch}">Whereabouts</th><th style="text-align:right;{bch}">Average</th><th style="text-align:right;{bch}">Highest</th></tr>')
        medal_bg = {1: '#FFF8E1', 2: '#F5F5F5', 3: '#FFF3E0'}
        for i, (gang, st) in enumerate(sorted(gs.items(), key=lambda x: x[1]['jv'], reverse=True), 1):
            bg = medal_bg.get(i, '#FFFFFF' if i % 2 == 0 else '#F8F9FA'); fw = 'bold' if i <= 3 else 'normal'
            h.append(f'<tr style="background:{bg};"><td style="text-align:center;{bc}font-weight:bold;">{i}</td><td style="{bc}font-weight:{fw};">{st["short"] or gang}</td><td style="{bc}">{st["pm"]}</td><td style="text-align:center;{bc}">{st["n"]}</td><td style="text-align:right;{bc}font-weight:{fw};">&pound;{st["jv"]:,.2f}</td><td style="text-align:right;{bc}">&pound;{st["wv"]:,.2f}</td><td style="text-align:right;{bc}">&pound;{st["avg"]:,.2f}</td><td style="text-align:right;{bc}">&pound;{st["max_val"]:,.2f}</td></tr>')
        h.append(f'<tr style="background:#E8EAF6;font-weight:bold;"><td style="{bch}"></td><td style="{bch}">TOTAL</td><td style="{bch}"></td><td style="text-align:center;{bch}">{s["total_jobs"]}</td><td style="text-align:right;{bch}">&pound;{s["total_jv"]:,.2f}</td><td style="text-align:right;{bch}">&pound;{s["total_wv"]:,.2f}</td><td style="text-align:right;{bch}">&pound;{s["avg_jv"]:,.2f}</td><td style="{bch}"></td></tr>')
        h.append('</table>')
        if self.daily_stats and s['n_days'] > 1:
            h.append('<p style="margin:14px 0 6px 0;font-weight:bold;color:#1B2A4A;font-size:13px;">Daily Summary</p>')
            h.append('<table cellpadding="5" cellspacing="0" style="border-collapse:collapse;font-size:12px;">')
            for i, day in enumerate(self.sorted_days):
                ds = self.daily_stats[day]['overall']; bg = '#F8F9FA' if i % 2 else '#FFFFFF'
                h.append(f'<tr style="background:{bg};"><td style="padding:4px 12px 4px 0;font-weight:bold;">{day}</td><td style="padding:4px 12px;">{ds["n"]} jobs</td><td style="padding:4px 12px;text-align:right;">&pound;{ds["jv"]:,.2f}</td><td style="padding:4px 12px;text-align:right;color:#888888;">&pound;{ds["avg"]:,.2f} avg</td></tr>')
            h.append('</table>')
        if self.finish_stats:
            h.append('<p style="margin:14px 0 6px 0;font-weight:bold;color:#1B2A4A;font-size:13px;">Finish Status</p>')
            h.append('<table cellpadding="4" cellspacing="0" style="border-collapse:collapse;font-size:12px;">')
            for status, fs in sorted(self.finish_stats.items(), key=lambda x: -x[1]['count']):
                bg = '#C6EFCE' if 'Complete' in status else '#FFF2CC' if 'Additional' in status else '#F8F9FA'
                h.append(f'<tr style="background:{bg};"><td style="padding:3px 12px;">{status}</td><td style="padding:3px 12px;text-align:center;">{fs["count"]}</td><td style="padding:3px 12px;text-align:right;">&pound;{fs["jv"]:,.2f}</td></tr>')
            h.append('</table>')
        if self.aborted_analysis is not None and len(self.aborted_analysis) > 0:
            n_ab = len(self.aborted_analysis)
            n_ov = int(self.aborted_analysis['in_whereabouts'].sum())
            jv_ov = round(float(self.aborted_analysis[self.aborted_analysis['in_whereabouts']]['jv'].sum()), 2)
            h.append(f'<p style="margin:14px 0 6px 0;font-weight:bold;color:#1B2A4A;font-size:13px;">Aborted Jobs: {n_ab} total | {n_ov} matched to whereabouts (&pound;{jv_ov:,.2f})</p>')
            h.append('<table cellpadding="4" cellspacing="0" style="border-collapse:collapse;font-size:12px;">')
            for reason, rs in sorted(self.aborted_by_reason.items(), key=lambda x: -x[1]['count']):
                bg = '#C6EFCE' if 'Complete' in reason else '#FFF2CC' if 'Planning' in reason else '#FFC7CE' if 'Access' in reason or 'Refused' in reason else '#F8F9FA'
                h.append(f'<tr style="background:{bg};"><td style="padding:3px 12px;">{reason}</td><td style="padding:3px 12px;text-align:center;">{rs["count"]}</td></tr>')
            h.append('</table>')
        h.append('<p style="color:#BBBBBB;font-size:10px;margin-top:18px;">Workbook attached - Whereabouts Values Analyzer v1.5</p>')
        h.append('</body></html>')
        return '\n'.join(h)


# ===================================================================
# REPORT (all values rounded to 2dp, all money cells use GF format)
# ===================================================================

class Report:
    def __init__(self, e): self.e = e; self.wb = None
    GF = f'{GBP}#,##0.00'

    def build(self, path, progress_cb=None):
        from openpyxl import Workbook; self.wb = Workbook()
        steps = [("League Table", self._league), ("PM Comparison", self._pm),
                 ("Value Analysis", self._value_analysis), ("By Work Type", self._work_type),
                 ("Daily Breakdown", lambda: self._daily() if len(self.e.daily_stats) > 1 else None),
                 ("Scheme Analysis", self._scheme),
                 ("Planned Overview", lambda: self._planned_overview() if self.e.is_planned_report else None),
                 ("Aborted Jobs", lambda: self._aborted() if self.e.aborted_analysis is not None else None),
                 ("All Jobs", self._all_jobs), ("Raw Data", self._raw)]
        for i, (name, fn) in enumerate(steps):
            if progress_cb: progress_cb(name, i + 1, len(steps))
            fn()
        self.wb.save(path)

    # -- styles --
    def _s(self):
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        return Font, PatternFill, Alignment, Border, Side
    def _hf(self):
        _, PF, *_ = self._s(); return PF(start_color=P['mid'], end_color=P['mid'], fill_type='solid')
    def _hft(self):
        F, *_ = self._s(); return F(name='Arial', size=11, bold=True, color='FFFFFF')
    def _wh(self, ws, row, hdrs):
        F, PF, A, B, S = self._s()
        for i, h in enumerate(hdrs, 1):
            c = ws.cell(row=row, column=i, value=h); c.font = self._hft(); c.fill = self._hf()
            c.alignment = A(horizontal='center', vertical='center', wrap_text=True)
            c.border = B(bottom=S(border_style='thin', color=P['bdr']))
    def _aw(self, ws, mn=12, mx=35):
        from openpyxl.cell.cell import MergedCell
        for col in ws.columns:
            ml = 0; cl = None
            for cell in col:
                if isinstance(cell, MergedCell): continue
                if cl is None: cl = cell.column_letter
                if cell.value: ml = max(ml, len(str(cell.value)))
            if cl: ws.column_dimensions[cl].width = min(max(ml + 3, mn), mx)
    def _title(self, ws, t, lc='N'):
        F, PF, A, *_ = self._s(); ws.merge_cells(f'A1:{lc}1'); c = ws['A1']; c.value = t
        c.font = F(name='Arial', size=15, bold=True, color='FFFFFF')
        c.fill = PF(start_color=P['dk'], end_color=P['dk'], fill_type='solid')
        c.alignment = A(horizontal='center', vertical='center'); ws.row_dimensions[1].height = 34
    def _sub(self, ws, t, lc='N'):
        F, PF, A, *_ = self._s(); ws.merge_cells(f'A2:{lc}2'); c = ws['A2']; c.value = t
        c.font = F(name='Arial', size=10, color='F4B942')
        c.fill = PF(start_color=P['dk'], end_color=P['dk'], fill_type='solid')
        c.alignment = A(horizontal='center', vertical='center'); ws.row_dimensions[2].height = 22
    def _rfill(self, i):
        _, PF, *_ = self._s(); c = P['alt'] if i % 2 else P['wh']
        return PF(start_color=c, end_color=c, fill_type='solid')
    def _dc(self, ws, r, c, v, fill=None, bold=False, fmt=None, color=None):
        F, *_ = self._s()
        from openpyxl.styles import Alignment as Al
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = F(name='Arial', size=10, bold=bold, color=color or '000000')
        cell.alignment = Al(horizontal='center', vertical='center')
        if fill: cell.fill = fill
        if fmt: cell.number_format = fmt
        return cell
    def _bar(self, val, total):
        pct = val / total * 100 if total else 0; f = int(pct / 5)
        return f"{chr(9608) * f}{chr(9617) * (20 - f)} {GBP}{val:,.2f} ({pct:.1f}%)"
    def _sec(self, ws, r, text):
        F, *_ = self._s(); ws.cell(row=r, column=1, value=text).font = F(name='Arial', size=13, bold=True, color=P['dk'])
    def _totals_row(self, ws, r, ncols):
        F, PF, *_ = self._s()
        for ci in range(1, ncols + 1):
            ws.cell(row=r, column=ci).font = F(name='Arial', size=10, bold=True)
            ws.cell(row=r, column=ci).fill = PF(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

    # -- 1. League Table --
    def _league(self):
        from openpyxl.chart import BarChart, Reference
        from openpyxl.utils import get_column_letter
        F, PF, A, *_ = self._s()
        ws = self.wb.active; ws.title = "League Table"; ws.sheet_view.showGridLines = False
        s = self.e.summary
        mode_label = "PLANNED REPORT" if s.get('is_planned_report') else "WHEREABOUTS VALUES v1.5"
        self._title(ws, f"{mode_label} -- {s['contract']} {s['region']}", 'K')
        period = f"{s['days'][0]} to {s['days'][-1]}" if s['n_days'] > 1 else (s['days'][0] if s['days'] else '')
        extras = []
        if s['n_duplicates'] > 0: extras.append(f"{s['n_duplicates']} shared jobs (values split)")
        if s['missing_wv_count'] > 0: extras.append(f"{GBP}{s['missing_wv_value']:,.2f} missing whereabouts")
        ext = '  |  '.join(extras)
        self._sub(ws, f"{period}  |  {s['n_pms']} PMs  |  {s['n_gangs']} Gangs  |  {s['total_jobs']} Jobs  |  {GBP}{s['total_jv']:,.2f} Total{'  |  ' + ext if ext else ''}", 'K')
        for i, (lbl, val, col) in enumerate([('TOTAL JOBS', s['total_jobs'], P['kpi_p']), ('JOB VALUE', f"{GBP}{s['total_jv']:,.2f}", P['kpi_g']),
            ('WHEREABOUTS', f"{GBP}{s['total_wv']:,.2f}", P['kpi_b']), ('AVERAGE', f"{GBP}{s['avg_jv']:,.2f}", P['kpi_a'])]):
            cx = 1 + i * 3
            ws.cell(row=4, column=cx, value=val).font = F(name='Arial', size=16, bold=True, color='2F5597')
            ws.cell(row=4, column=cx).fill = PF(start_color=col, end_color=col, fill_type='solid')
            ws.cell(row=4, column=cx).alignment = A(horizontal='center', vertical='center')
            ws.cell(row=5, column=cx, value=lbl).font = F(name='Arial', size=9, bold=True, color='404040')
            ws.cell(row=5, column=cx).fill = PF(start_color=col, end_color=col, fill_type='solid')
            ws.cell(row=5, column=cx).alignment = A(horizontal='center')
        hdrs = ['Rank', 'Gang', 'PM', 'Jobs', 'Job Value', 'Whereabouts Value', 'Average Job Value', 'Capital', 'Defects', 'Highest Job', 'Share of Total']
        self._wh(ws, 7, hdrs)
        ranked = sorted(self.e.gang_stats.items(), key=lambda x: x[1]['jv'], reverse=True)
        mf = [PF(start_color=P['gold'], end_color=P['gold'], fill_type='solid'), PF(start_color=P['silver'], end_color=P['silver'], fill_type='solid'), PF(start_color=P['bronze'], end_color=P['bronze'], fill_type='solid')]
        for idx, (gang, st) in enumerate(ranked, 1):
            r = 7 + idx; rf = mf[idx-1] if idx <= 3 else self._rfill(idx); b = idx <= 3
            self._dc(ws,r,1,idx,fill=rf,bold=b); self._dc(ws,r,2,gang,fill=rf,bold=b); self._dc(ws,r,3,st['pm'],fill=rf,bold=b)
            self._dc(ws,r,4,st['n'],fill=rf,bold=b); self._dc(ws,r,5,st['jv'],fill=rf,bold=b,fmt=self.GF)
            self._dc(ws,r,6,st['wv'],fill=rf,bold=b,fmt=self.GF); self._dc(ws,r,7,st['avg'],fill=rf,bold=b,fmt=self.GF)
            self._dc(ws,r,8,st['cap'],fill=rf,bold=b); self._dc(ws,r,9,st['defects'],fill=rf,bold=b)
            self._dc(ws,r,10,st['max_val'],fill=rf,bold=b,fmt=self.GF); self._dc(ws,r,11,self._bar(st['jv'],s['total_jv']),fill=rf)
        last = 7 + len(ranked); tr = last + 1; self._dc(ws,tr,2,'TOTALS',bold=True)
        for ci in [4,5,6,8,9]:
            self._dc(ws,tr,ci,f'=SUM({get_column_letter(ci)}8:{get_column_letter(ci)}{last})',bold=True,fmt=self.GF if ci in [5,6] else None)
        self._totals_row(ws, tr, len(hdrs))
        ws.freeze_panes = 'A8'; ws.auto_filter.ref = f"A7:K{last}"
        try:
            ch = BarChart(); ch.title = "Job Value by Gang"; ch.style = 10; ch.height = 12; ch.width = 22
            ch.add_data(Reference(ws,min_col=5,min_row=7,max_row=last),titles_from_data=True)
            ch.set_categories(Reference(ws,min_col=2,min_row=8,max_row=last)); ws.add_chart(ch,f"A{tr+2}")
        except: pass
        self._aw(ws,mx=50); ws.column_dimensions['B'].width = 36; ws.column_dimensions['K'].width = 40

    # -- 2. PM Comparison --
    def _pm(self):
        F, PF, *_ = self._s()
        ws = self.wb.create_sheet("PM Comparison"); ws.sheet_view.showGridLines = False
        self._title(ws,"PM COMPARISON -- HEAD-TO-HEAD",'I'); self._sub(ws,"Ranked by total job value",'I')
        self._wh(ws,4,['PM','Teams','Jobs','Job Value','Whereabouts Value','Average Job Value','Capital','Defects','Share'])
        r = 5; s = self.e.summary
        for pm, ps in sorted(self.e.pm_stats.items(), key=lambda x: x[1]['jv'], reverse=True):
            rf = self._rfill(r)
            self._dc(ws,r,1,pm,fill=rf,bold=True); self._dc(ws,r,2,ps['num_teams'],fill=rf); self._dc(ws,r,3,ps['n'],fill=rf)
            self._dc(ws,r,4,ps['jv'],fill=rf,fmt=self.GF); self._dc(ws,r,5,ps['wv'],fill=rf,fmt=self.GF)
            self._dc(ws,r,6,ps['avg'],fill=rf,fmt=self.GF); self._dc(ws,r,7,ps['cap'],fill=rf)
            self._dc(ws,r,8,ps['defects'],fill=rf); self._dc(ws,r,9,self._bar(ps['jv'],s['total_jv']),fill=rf); r += 1
        r += 1
        for pm, ps in sorted(self.e.pm_stats.items(), key=lambda x: x[1]['jv'], reverse=True):
            self._sec(ws,r,f"{pm} -- Team Breakdown"); r += 1
            self._wh(ws,r,['Team','Jobs','Job Value','Whereabouts Value','Average Job Value']); r += 1
            for tb in sorted(ps['breakdown'], key=lambda x: x['jv'], reverse=True):
                rf = self._rfill(r); self._dc(ws,r,1,tb['short'],fill=rf,bold=True); self._dc(ws,r,2,tb['n'],fill=rf)
                self._dc(ws,r,3,tb['jv'],fill=rf,fmt=self.GF); self._dc(ws,r,4,tb['wv'],fill=rf,fmt=self.GF)
                self._dc(ws,r,5,tb['avg'],fill=rf,fmt=self.GF); r += 1
            r += 1
        ws.freeze_panes = 'A5'; self._aw(ws,mx=50); ws.column_dimensions['I'].width = 40

    # -- 3. Value Analysis (the big one) --
    def _value_analysis(self):
        F, PF, A, *_ = self._s()
        ws = self.wb.create_sheet("Value Analysis"); ws.sheet_view.showGridLines = False
        self._title(ws,"VALUE ANALYSIS -- DEEP DIVE",'L')
        self._sub(ws,"Run rate | Variance | Missing whereabouts | Distribution | Every job by day",'L')

        # Gang analysis
        self._wh(ws,4,['Gang','PM','Jobs','Total Value','Average','Days Active','Daily Run Rate','Variance from Average','Variance %','Whereabouts Value','Whereabouts Gap','Whereabouts %'])
        r = 5
        for gang, va in sorted(self.e.value_analysis.items(), key=lambda x: -x[1]['daily_rate']):
            rf = self._rfill(r); vc = '006100' if va['variance'] >= 0 else '9C0006'
            self._dc(ws,r,1,va['short'] or gang,fill=rf,bold=True); self._dc(ws,r,2,va['pm'],fill=rf)
            self._dc(ws,r,3,va['n'],fill=rf); self._dc(ws,r,4,va['jv'],fill=rf,fmt=self.GF)
            self._dc(ws,r,5,va['avg'],fill=rf,fmt=self.GF); self._dc(ws,r,6,va['days_active'],fill=rf)
            self._dc(ws,r,7,va['daily_rate'],fill=rf,fmt=self.GF)
            self._dc(ws,r,8,va['variance'],fill=rf,fmt=self.GF,color=vc)
            self._dc(ws,r,9,va['variance_pct'],fill=rf,fmt='0.0"%"')
            self._dc(ws,r,10,va['wv'],fill=rf,fmt=self.GF); self._dc(ws,r,11,va['wv_gap'],fill=rf,fmt=self.GF)
            self._dc(ws,r,12,va['wv_ratio'],fill=rf,fmt='0.0"%"'); r += 1

        # Missing whereabouts
        if self.e.missing_wv_by_gang:
            r += 2; s = self.e.summary; warn_fill = PF(start_color=P['warn'],end_color=P['warn'],fill_type='solid')
            self._sec(ws,r,f"MISSING WHEREABOUTS -- {s['missing_wv_count']} jobs | {GBP}{s['missing_wv_value']:,.2f} with ZERO whereabouts"); r += 1
            self._wh(ws,r,['Gang','PM','Missing Jobs','Total Jobs','% Missing','Missing Job Value','Total Gang Value','% of Gang Value']); r += 1
            for gang, mg in sorted(self.e.missing_wv_by_gang.items(), key=lambda x: -x[1]['missing_jv']):
                rf = warn_fill if mg['missing_pct'] > 50 else self._rfill(r)
                self._dc(ws,r,1,mg['short'] or gang,fill=rf,bold=True); self._dc(ws,r,2,mg['pm'],fill=rf)
                self._dc(ws,r,3,mg['missing_count'],fill=rf); self._dc(ws,r,4,mg['total_count'],fill=rf)
                self._dc(ws,r,5,mg['missing_pct'],fill=rf,fmt='0.0"%"'); self._dc(ws,r,6,mg['missing_jv'],fill=rf,fmt=self.GF)
                self._dc(ws,r,7,mg['gang_jv'],fill=rf,fmt=self.GF); self._dc(ws,r,8,mg['missing_jv_pct'],fill=rf,fmt='0.0"%"'); r += 1
            r += 2; self._sec(ws,r,"MISSING WHEREABOUTS -- INDIVIDUAL JOBS (Top 20 by Value)"); r += 1
            mj_cols = [c for c in ['Job ID','Client Ref 1','Gang','_pm','_scheme','Work Type','Job Value','_day'] if c in self.e.missing_wv.columns]
            cm = {'Job ID':'Job ID','Client Ref 1':'Client Ref','Gang':'Gang','_pm':'PM','_scheme':'Scheme','Work Type':'Work Type','Job Value':'Job Value','_day':'Date'}
            self._wh(ws,r,[cm.get(c,c) for c in mj_cols]); r += 1
            for _, rd in self.e.missing_wv.nlargest(20,'Job Value').iterrows():
                rf = self._rfill(r)
                for ci, col in enumerate(mj_cols, 1):
                    val = rd[col]
                    if pd.isna(val): val = ''
                    elif isinstance(val, float): val = round(val, 2)
                    cell = self._dc(ws,r,ci,val,fill=rf)
                    if col == 'Job Value': cell.number_format = self.GF
                r += 1

        # Shared/duplicate jobs
        if self.e.duplicates:
            r += 2
            self._sec(ws,r,f"SHARED JOBS -- {len(self.e.duplicates)} jobs on multiple gangs | Value assigned to first gang, others show {GBP}0.00"); r += 1
            dup_hdrs = ['Job ID','Client Ref','Scheme','Work Type','Date','Job Value','Whereabouts','Value Assigned To','Also On','Gangs']
            self._wh(ws,r,dup_hdrs); r += 1
            _, PFw, *_ = self._s()
            warn_fill = PFw(start_color=P['warn'],end_color=P['warn'],fill_type='solid')
            for d in sorted(self.e.duplicates, key=lambda x: -x['orig_val']):
                rf = warn_fill
                others = [g for g in d['gang_shorts'] if g != d['owner']]
                self._dc(ws,r,1,d['job_id'],fill=rf,bold=True)
                self._dc(ws,r,2,d['client_ref'],fill=rf)
                self._dc(ws,r,3,d['scheme'],fill=rf)
                self._dc(ws,r,4,d['work_type'],fill=rf)
                self._dc(ws,r,5,d['day'],fill=rf)
                self._dc(ws,r,6,d['orig_val'],fill=rf,fmt=self.GF)
                self._dc(ws,r,7,d['orig_wv'],fill=rf,fmt=self.GF)
                self._dc(ws,r,8,d['owner'],fill=rf,bold=True)
                self._dc(ws,r,9,', '.join(others),fill=rf)
                self._dc(ws,r,10,d['count'],fill=rf)
                r += 1
            self._totals_row(ws, r, len(dup_hdrs))
            self._dc(ws,r,1,f'TOTAL ({len(self.e.duplicates)} shared jobs)',bold=True)
            self._dc(ws,r,6,self.e.duplicate_true_value,bold=True,fmt=self.GF)
            r += 1

        # Value distribution
        r += 2; self._sec(ws,r,"VALUE DISTRIBUTION"); r += 1
        self._wh(ws,r,['Value Band','Jobs','% of Jobs','Total Value','% of Total Value','Average Value']); r += 1
        for band, bd in self.e.value_bands.items():
            rf = self._rfill(r)
            self._dc(ws,r,1,f"{GBP}{band}",fill=rf,bold=True); self._dc(ws,r,2,bd['count'],fill=rf)
            self._dc(ws,r,3,bd['pct_count'],fill=rf,fmt='0.0"%"'); self._dc(ws,r,4,bd['jv'],fill=rf,fmt=self.GF)
            self._dc(ws,r,5,bd['pct_value'],fill=rf,fmt='0.0"%"'); self._dc(ws,r,6,bd['avg'],fill=rf,fmt=self.GF); r += 1

        # Finish status
        r += 2; self._sec(ws,r,"FINISH STATUS BREAKDOWN"); r += 1
        self._wh(ws,r,['Status','Jobs','% of Total','Job Value','Whereabouts Value']); r += 1
        for status, fs in sorted(self.e.finish_stats.items(), key=lambda x: -x[1]['count']):
            rf = self._rfill(r)
            if 'Complete' in status: _, PFl, *_ = self._s(); rf = PFl(start_color=P['g'],end_color=P['g'],fill_type='solid')
            elif 'Additional' in status: _, PFl, *_ = self._s(); rf = PFl(start_color=P['a'],end_color=P['a'],fill_type='solid')
            self._dc(ws,r,1,status,fill=rf,bold=True); self._dc(ws,r,2,fs['count'],fill=rf)
            self._dc(ws,r,3,fs['pct'],fill=rf,fmt='0.0"%"'); self._dc(ws,r,4,fs['jv'],fill=rf,fmt=self.GF)
            self._dc(ws,r,5,fs['wv'],fill=rf,fmt=self.GF); r += 1

        # Top 10
        r += 2; self._sec(ws,r,"TOP 10 HIGHEST VALUE JOBS"); r += 1
        top = self.e.top_jobs
        if top is not None and not top.empty:
            cm = {'Job ID':'Job ID','Client Ref 1':'Client Ref','Gang':'Gang','_pm':'PM','_scheme':'Scheme','Work Type':'Work Type','Job Value':'Job Value','Whereabouts Value':'Whereabouts','Address':'Address','_day':'Date'}
            self._wh(ws,r,[cm.get(c,c) for c in top.columns]); r += 1
            for _, rd in top.iterrows():
                rf = self._rfill(r)
                for ci, col in enumerate(top.columns, 1):
                    val = rd[col]
                    if pd.isna(val): val = ''
                    elif isinstance(val, float): val = round(val, 2)
                    cell = self._dc(ws,r,ci,val,fill=rf)
                    if col in ('Job Value','Whereabouts Value'): cell.number_format = self.GF
                r += 1

        # EVERY JOB BY DAY
        r += 2; self._sec(ws,r,"ALL JOBS BY DAY -- EVERY JOB WITH VALUE"); r += 1
        job_cols = [c for c in ['_day','Gang','_pm','Job ID','Client Ref 1','Work Type','Job Value','Whereabouts Value','Address','Finish Status'] if c in self.e.data.columns]
        job_hdrs = ['Date','Gang','PM','Job ID','Client Ref','Work Type','Job Value','Whereabouts Value','Address','Finish Status'][:len(job_cols)]
        self._wh(ws,r,job_hdrs); r += 1
        df = self.e.data.copy()
        jv_ci = job_cols.index('Job Value') + 1 if 'Job Value' in job_cols else None
        wv_ci = job_cols.index('Whereabouts Value') + 1 if 'Whereabouts Value' in job_cols else None
        for day in self.e.sorted_days:
            dd = df[df['_day'] == day].sort_values(['Gang','Job Value'], ascending=[True, False])
            for _, rd in dd.iterrows():
                rf = self._rfill(r)
                for ci, col in enumerate(job_cols, 1):
                    val = rd[col]
                    if pd.isna(val): val = ''
                    elif isinstance(val, float): val = round(val, 2)
                    cell = self._dc(ws,r,ci,val,fill=rf)
                    if col in ('Job Value','Whereabouts Value'): cell.number_format = self.GF
                r += 1
            ds = self.e.daily_stats[day]['overall']
            self._totals_row(ws, r, len(job_hdrs))
            self._dc(ws,r,1,f'{day} TOTAL ({len(dd)} jobs)',bold=True)
            if jv_ci: self._dc(ws,r,jv_ci,ds['jv'],bold=True,fmt=self.GF)
            if wv_ci: self._dc(ws,r,wv_ci,ds['wv'],bold=True,fmt=self.GF)
            r += 1
        # Grand total
        for ci in range(1, len(job_hdrs) + 1):
            c = ws.cell(row=r, column=ci)
            c.font = F(name='Arial', size=11, bold=True, color='FFFFFF')
            c.fill = PF(start_color=P['dk'], end_color=P['dk'], fill_type='solid')
        self._dc(ws,r,1,f'GRAND TOTAL ({len(df)} jobs)',bold=True,color='FFFFFF')
        ws.cell(row=r,column=1).font = F(name='Arial',size=11,bold=True,color='FFFFFF')
        if jv_ci:
            c = self._dc(ws,r,jv_ci,self.e.summary['total_jv'],bold=True,fmt=self.GF,color='FFFFFF')
            c.font = F(name='Arial',size=11,bold=True,color='FFFFFF')
        if wv_ci:
            c = self._dc(ws,r,wv_ci,self.e.summary['total_wv'],bold=True,fmt=self.GF,color='FFFFFF')
            c.font = F(name='Arial',size=11,bold=True,color='FFFFFF')

        ws.freeze_panes = 'A5'; self._aw(ws,mx=35); ws.column_dimensions['A'].width = 30; ws.column_dimensions['B'].width = 36

    # -- 4. By Work Type --
    def _work_type(self):
        ws = self.wb.create_sheet("By Work Type"); ws.sheet_view.showGridLines = False
        self._title(ws,"BREAKDOWN BY WORK TYPE",'F'); self._sub(ws,"Each gang split by Capital Works vs Defects",'F')
        self._wh(ws,4,['Gang','Work Type','Jobs','Job Value','Whereabouts Value','Average Job Value']); r = 5
        for (gang,wt), st in sorted(self.e.wt_stats.items(), key=lambda x: (x[0][0], -x[1]['jv'])):
            rf = self._rfill(r); self._dc(ws,r,1,gang,fill=rf,bold=True); self._dc(ws,r,2,wt,fill=rf)
            self._dc(ws,r,3,st['n'],fill=rf); self._dc(ws,r,4,st['jv'],fill=rf,fmt=self.GF)
            self._dc(ws,r,5,st['wv'],fill=rf,fmt=self.GF); self._dc(ws,r,6,st['avg'],fill=rf,fmt=self.GF); r += 1
        ws.auto_filter.ref = f"A4:F{r-1}"; ws.freeze_panes = 'A5'; self._aw(ws,mx=40); ws.column_dimensions['A'].width = 36

    # -- 5. Daily Breakdown --
    def _daily(self):
        F, PF, *_ = self._s()
        ws = self.wb.create_sheet("Daily Breakdown"); ws.sheet_view.showGridLines = False
        self._title(ws,"DAILY BREAKDOWN",'G'); self._sub(ws,"Day-by-day gang values with subtotals",'G')
        self._wh(ws,4,['Date','Gang','PM','Jobs','Job Value','Whereabouts Value','Average Job Value']); r = 5
        for day in self.e.sorted_days:
            for gang, st in sorted(self.e.daily_stats[day]['gangs'].items(), key=lambda x: -x[1]['jv']):
                pm = self.e.gang_stats.get(gang, {}).get('pm', ''); rf = self._rfill(r)
                self._dc(ws,r,1,day,fill=rf); self._dc(ws,r,2,gang,fill=rf,bold=True); self._dc(ws,r,3,pm,fill=rf)
                self._dc(ws,r,4,st['n'],fill=rf); self._dc(ws,r,5,st['jv'],fill=rf,fmt=self.GF)
                self._dc(ws,r,6,st['wv'],fill=rf,fmt=self.GF); self._dc(ws,r,7,st['avg'],fill=rf,fmt=self.GF); r += 1
            ds = self.e.daily_stats[day]['overall']; self._totals_row(ws,r,7)
            self._dc(ws,r,1,f'{day} TOTAL',bold=True); self._dc(ws,r,4,ds['n'],bold=True)
            self._dc(ws,r,5,ds['jv'],bold=True,fmt=self.GF); self._dc(ws,r,6,ds['wv'],bold=True,fmt=self.GF)
            self._dc(ws,r,7,ds['avg'],bold=True,fmt=self.GF); r += 1
        ws.auto_filter.ref = f"A4:G{r-1}"; ws.freeze_panes = 'A5'; self._aw(ws,mx=40); ws.column_dimensions['B'].width = 36

    # -- 6. Scheme Analysis --
    def _scheme(self):
        ws = self.wb.create_sheet("Scheme Analysis"); ws.sheet_view.showGridLines = False
        self._title(ws,"SCHEME PERFORMANCE",'H'); self._sub(ws,"Job values by contract/scheme",'H')
        self._wh(ws,4,['Scheme','Jobs','Job Value','Whereabouts Value','Average Job Value','Gangs','Capital','Defects']); r = 5
        for sch, st in sorted(self.e.scheme_stats.items(), key=lambda x: -x[1]['jv']):
            rf = self._rfill(r); self._dc(ws,r,1,sch,fill=rf,bold=True); self._dc(ws,r,2,st['n'],fill=rf)
            self._dc(ws,r,3,st['jv'],fill=rf,fmt=self.GF); self._dc(ws,r,4,st['wv'],fill=rf,fmt=self.GF)
            self._dc(ws,r,5,st['avg'],fill=rf,fmt=self.GF); self._dc(ws,r,6,st['teams'],fill=rf)
            self._dc(ws,r,7,st['cap'],fill=rf); self._dc(ws,r,8,st['defects'],fill=rf); r += 1
        ws.auto_filter.ref = f"A4:H{r-1}"; ws.freeze_panes = 'A5'; self._aw(ws,mx=40); ws.column_dimensions['A'].width = 30

    # -- 7. Planned Overview --
    def _planned_overview(self):
        F, PF, A, *_ = self._s()
        ws = self.wb.create_sheet("Planned Overview"); ws.sheet_view.showGridLines = False
        ps = self.e.planned_stats
        if not ps: return
        s = self.e.summary
        self._title(ws, f"PLANNED REPORT OVERVIEW -- {s['contract']} {s['region']}", 'N')
        self._sub(ws, f"Programmed vs Forecast | {s['total_jobs']} total jobs | Measure rate: {ps['measure_rate']:.1f}%", 'N')

        # KPI row
        for i, (lbl, val, col) in enumerate([
            ('PLANNED VALUE', f"{GBP}{ps['total_planned']:,.2f}", P['kpi_g']),
            ('MEASURED VALUE', f"{GBP}{ps['total_measured']:,.2f}", P['kpi_b']),
            ('REMAINING', f"{GBP}{ps['total_remaining']:,.2f}", P['kpi_a']),
            ('MEASURE RATE', f"{ps['measure_rate']:.1f}%", P['kpi_p']),
        ]):
            cx = 1 + i * 3
            ws.cell(row=4, column=cx, value=val).font = F(name='Arial', size=16, bold=True, color='2F5597')
            ws.cell(row=4, column=cx).fill = PF(start_color=col, end_color=col, fill_type='solid')
            ws.cell(row=4, column=cx).alignment = A(horizontal='center', vertical='center')
            ws.cell(row=5, column=cx, value=lbl).font = F(name='Arial', size=9, bold=True, color='404040')
            ws.cell(row=5, column=cx).fill = PF(start_color=col, end_color=col, fill_type='solid')
            ws.cell(row=5, column=cx).alignment = A(horizontal='center')

        r = 7

        # Job Status breakdown
        if 'status_breakdown' in ps:
            self._sec(ws, r, "JOB STATUS BREAKDOWN"); r += 1
            self._wh(ws, r, ['Job Status', 'Jobs', 'Planned Value', 'Measured Value', 'Remaining Value', '% Measured']); r += 1
            for status, sb in sorted(ps['status_breakdown'].items(), key=lambda x: -x[1]['planned']):
                rf = self._rfill(r)
                if status == 'Complete': rf = PF(start_color=P['g'], end_color=P['g'], fill_type='solid')
                elif status == 'Site Clear': rf = PF(start_color=P['b'], end_color=P['b'], fill_type='solid')
                elif status == 'In Progress': rf = PF(start_color=P['a'], end_color=P['a'], fill_type='solid')
                mr = round(sb['measured'] / sb['planned'] * 100, 1) if sb['planned'] > 0 else 0
                self._dc(ws,r,1,status,fill=rf,bold=True); self._dc(ws,r,2,sb['count'],fill=rf)
                self._dc(ws,r,3,sb['planned'],fill=rf,fmt=self.GF); self._dc(ws,r,4,sb['measured'],fill=rf,fmt=self.GF)
                self._dc(ws,r,5,sb['remaining'],fill=rf,fmt=self.GF); self._dc(ws,r,6,mr,fill=rf,fmt='0.0"%"'); r += 1
            self._totals_row(ws, r, 6)
            self._dc(ws,r,1,'TOTAL',bold=True); self._dc(ws,r,2,s['total_jobs'],bold=True)
            self._dc(ws,r,3,ps['total_planned'],bold=True,fmt=self.GF); self._dc(ws,r,4,ps['total_measured'],bold=True,fmt=self.GF)
            self._dc(ws,r,5,ps['total_remaining'],bold=True,fmt=self.GF)
            self._dc(ws,r,6,ps['measure_rate'],bold=True,fmt='0.0"%"'); r += 2

        # Key metrics summary
        self._sec(ws, r, "KEY METRICS"); r += 1
        self._wh(ws, r, ['Metric', 'Count', 'Value']); r += 1
        metrics = [
            ('Jobs Site Cleared', ps['n_site_cleared'], ''),
            ('Jobs Started', ps['n_started'], ''),
            ('Measures Logged', ps['n_measures_logged'], f"{GBP}{ps['total_prev_measures']:,.2f}"),
            ('Missing Measures', ps['n_missing_measures'], ''),
            ('Aborted', ps['n_aborted'], ''),
            ('Unscheduled', ps['n_unscheduled'], ''),
            ('Cost Variations', ps['n_cv'], f"{GBP}{ps['total_cv_value']:,.2f}"),
            ('On Hold', ps['n_on_hold'], ''),
            ('Ready To Invoice', ps['n_ready_invoice'], ''),
            ('Pending Measures', '', f"{GBP}{ps['total_pending_val']:,.2f}"),
        ]
        for label, count, val in metrics:
            rf = self._rfill(r)
            if 'Missing' in label or 'Aborted' in label:
                rf = PF(start_color=P['r'], end_color=P['r'], fill_type='solid')
            elif 'Cleared' in label or 'Invoice' in label:
                rf = PF(start_color=P['g'], end_color=P['g'], fill_type='solid')
            self._dc(ws,r,1,label,fill=rf,bold=True)
            self._dc(ws,r,2,count if count != '' else '',fill=rf)
            if val and val.startswith(GBP):
                num = float(val.replace(GBP,'').replace(',',''))
                self._dc(ws,r,3,num,fill=rf,fmt=self.GF)
            else:
                self._dc(ws,r,3,val,fill=rf)
            r += 1
        r += 1

        # Gang Planned vs Measured
        self._sec(ws, r, "GANG -- PLANNED vs MEASURED"); r += 1
        self._wh(ws, r, ['Gang', 'PM', 'Jobs', 'Planned Value', 'Measured Value', 'Remaining',
                          'Measure Rate', 'Site Cleared', 'Started', 'Missing Measures', 'Aborted', 'CVs', 'CV Value']); r += 1
        gp = ps.get('gang_planned', {})
        gs = self.e.gang_stats
        for gang in sorted(gp.keys(), key=lambda g: -gp[g]['planned']):
            rf = self._rfill(r)
            gps = gp[gang]; gst = gs.get(gang, {})
            pm = gst.get('pm', ''); short = gst.get('short', gang); n = gst.get('n', 0)
            if gps['measure_rate'] < 20:
                rf = PF(start_color=P['r'], end_color=P['r'], fill_type='solid')
            elif gps['measure_rate'] < 50:
                rf = PF(start_color=P['a'], end_color=P['a'], fill_type='solid')
            self._dc(ws,r,1,short,fill=rf,bold=True); self._dc(ws,r,2,pm,fill=rf)
            self._dc(ws,r,3,n,fill=rf); self._dc(ws,r,4,gps['planned'],fill=rf,fmt=self.GF)
            self._dc(ws,r,5,gps['measured'],fill=rf,fmt=self.GF); self._dc(ws,r,6,gps['remaining'],fill=rf,fmt=self.GF)
            self._dc(ws,r,7,gps['measure_rate'],fill=rf,fmt='0.0"%"')
            self._dc(ws,r,8,gps['n_site_cleared'],fill=rf); self._dc(ws,r,9,gps['n_started'],fill=rf)
            self._dc(ws,r,10,gps['n_missing_measures'],fill=rf); self._dc(ws,r,11,gps['n_aborted'],fill=rf)
            self._dc(ws,r,12,gps['n_cv'],fill=rf); self._dc(ws,r,13,gps['cv_value'],fill=rf,fmt=self.GF)
            r += 1
        self._totals_row(ws, r, 13)
        self._dc(ws,r,1,'TOTAL',bold=True); self._dc(ws,r,3,s['total_jobs'],bold=True)
        self._dc(ws,r,4,ps['total_planned'],bold=True,fmt=self.GF); self._dc(ws,r,5,ps['total_measured'],bold=True,fmt=self.GF)
        self._dc(ws,r,6,ps['total_remaining'],bold=True,fmt=self.GF)
        self._dc(ws,r,7,ps['measure_rate'],bold=True,fmt='0.0"%"')

        ws.freeze_panes = 'A7'; self._aw(ws, mx=35)
        ws.column_dimensions['A'].width = 30; ws.column_dimensions['B'].width = 20

    # -- 8. Aborted Jobs --
    def _aborted(self):
        F, PF, A, *_ = self._s()
        ws = self.wb.create_sheet("Aborted Jobs"); ws.sheet_view.showGridLines = False
        ab = self.e.aborted_analysis
        if ab is None or ab.empty: return

        n_ab = len(ab); jv_ab = round(float(ab['jv'].sum()), 2); n_in_wv = int(ab['in_whereabouts'].sum())
        n_no_val = n_ab - n_in_wv

        # Calculate estimated value for jobs not in whereabouts using avg defect rate
        defect_data = self.e.data[self.e.data['Work Type'].str.contains('Defect', case=False, na=False)]
        avg_defect = round(float(defect_data['_orig_jv'].mean()), 2) if len(defect_data) > 0 else 0
        est_total = round(jv_ab + (n_no_val * avg_defect), 2) if avg_defect > 0 else jv_ab

        # Daily breakdown of aborts
        abort_by_day = {}
        if 'aborted_date' in ab.columns:
            for _, row in ab.iterrows():
                d = row['aborted_date']
                day_str = d.strftime('%d/%m') if hasattr(d, 'strftime') else str(d)[:5] if d else 'Unknown'
                if day_str not in abort_by_day:
                    abort_by_day[day_str] = {'count': 0, 'jv': 0}
                abort_by_day[day_str]['count'] += 1
                abort_by_day[day_str]['jv'] = round(abort_by_day[day_str]['jv'] + row['jv'], 2)

        # Top abort reason
        top_reason = max(self.e.aborted_by_reason.items(), key=lambda x: x[1]['count'])[0] if self.e.aborted_by_reason else ''
        top_reason_count = self.e.aborted_by_reason.get(top_reason, {}).get('count', 0)
        abort_rate = round(n_ab / (n_ab + self.e.summary['total_jobs']) * 100, 1) if (n_ab + self.e.summary['total_jobs']) > 0 else 0

        self._title(ws, f"ABORTED JOBS -- {n_ab} aborted | {n_in_wv} matched to whereabouts | {n_no_val} unmatched", 'L')
        self._sub(ws, f"Known value: {GBP}{jv_ab:,.2f} ({n_in_wv} jobs) | Estimated total: {GBP}{est_total:,.2f} (avg defect rate {GBP}{avg_defect:,.2f})", 'L')

        # KPI scoreboard - row 4 and 5
        for i, (lbl, val, col) in enumerate([
            ('TOTAL ABORTED', n_ab, P['r']),
            ('KNOWN VALUE', f"{GBP}{jv_ab:,.2f}", P['kpi_a']),
            ('ESTIMATED TOTAL', f"{GBP}{est_total:,.2f}", P['kpi_p']),
            ('ABORT RATE', f"{abort_rate}%", P['kpi_b']),
        ]):
            cx = 1 + i * 3
            ws.cell(row=4, column=cx, value=val).font = F(name='Arial', size=16, bold=True, color='2F5597')
            ws.cell(row=4, column=cx).fill = PF(start_color=col, end_color=col, fill_type='solid')
            ws.cell(row=4, column=cx).alignment = A(horizontal='center', vertical='center')
            ws.cell(row=5, column=cx, value=lbl).font = F(name='Arial', size=9, bold=True, color='404040')
            ws.cell(row=5, column=cx).fill = PF(start_color=col, end_color=col, fill_type='solid')
            ws.cell(row=5, column=cx).alignment = A(horizontal='center')

        r = 7

        # Daily abort cost
        if abort_by_day:
            self._sec(ws, r, "DAILY ABORT COST"); r += 1
            self._wh(ws, r, ['Date', 'Jobs Aborted', 'Known Value Lost', 'Estimated Value Lost']); r += 1
            for day in sorted(abort_by_day.keys()):
                dd = abort_by_day[day]
                rf = self._rfill(r)
                n_unmatched_day = dd['count'] - len(ab[(ab['aborted_date'].apply(lambda x: x.strftime('%d/%m') if hasattr(x, 'strftime') else '') == day) & (ab['in_whereabouts'])])
                est_day = round(dd['jv'] + (n_unmatched_day * avg_defect), 2)
                self._dc(ws,r,1,day,fill=rf,bold=True)
                self._dc(ws,r,2,dd['count'],fill=rf)
                if dd['jv'] > 0:
                    self._dc(ws,r,3,dd['jv'],fill=rf,fmt=self.GF)
                else:
                    self._dc(ws,r,3,'N/A',fill=rf)
                self._dc(ws,r,4,est_day,fill=rf,fmt=self.GF)
                r += 1
            r += 1

        # Section: Summary by abort reason
        self._sec(ws, r, "ABORT REASON SUMMARY"); r += 1
        self._wh(ws, r, ['Abort Reason', 'Count', 'Job Value (from Whereabouts)', 'Also in Whereabouts', '% of Aborts']); r += 1
        for reason, rs in sorted(self.e.aborted_by_reason.items(), key=lambda x: -x[1]['count']):
            rf = self._rfill(r)
            pct = round(rs['count'] / n_ab * 100, 1) if n_ab else 0
            self._dc(ws,r,1,reason,fill=rf,bold=True); self._dc(ws,r,2,rs['count'],fill=rf)
            if rs['in_wv'] > 0:
                self._dc(ws,r,3,rs['jv'],fill=rf,fmt=self.GF)
            else:
                self._dc(ws,r,3,'N/A',fill=rf)
            self._dc(ws,r,4,rs['in_wv'],fill=rf)
            self._dc(ws,r,5,pct,fill=rf,fmt='0.0"%"'); r += 1
        self._totals_row(ws, r, 5)
        self._dc(ws,r,1,'TOTAL',bold=True); self._dc(ws,r,2,n_ab,bold=True)
        self._dc(ws,r,3,jv_ab,bold=True,fmt=self.GF); self._dc(ws,r,4,n_in_wv,bold=True)
        r += 2

        # Section 2: Summary by gang
        self._sec(ws, r, "ABORTS BY GANG"); r += 1
        self._wh(ws, r, ['Gang', 'Aborted', 'Job Value (from Whereabouts)', 'In Whereabouts', 'Top Reason']); r += 1
        for gang, gs in sorted(self.e.aborted_by_gang.items(), key=lambda x: -x[1]['count']):
            rf = self._rfill(r)
            top_reason = max(gs['reasons'], key=gs['reasons'].get) if gs['reasons'] else ''
            self._dc(ws,r,1,gs['gang_short'],fill=rf,bold=True); self._dc(ws,r,2,gs['count'],fill=rf)
            if gs['in_wv'] > 0:
                self._dc(ws,r,3,gs['jv'],fill=rf,fmt=self.GF)
            else:
                self._dc(ws,r,3,'N/A',fill=rf)
            self._dc(ws,r,4,gs['in_wv'],fill=rf)
            self._dc(ws,r,5,top_reason,fill=rf); r += 1
        r += 2

        # Section 3: Jobs in BOTH aborted and whereabouts
        overlap = ab[ab['in_whereabouts']].sort_values('jv', ascending=False)
        if len(overlap) > 0:
            self._sec(ws, r, f"JOBS IN BOTH FILES -- {len(overlap)} aborted jobs also in whereabouts ({GBP}{round(float(overlap['jv'].sum()), 2):,.2f} value)"); r += 1
            self._wh(ws, r, ['Job ID', 'Gang', 'Abort Reason', 'Job Value', 'Whereabouts Value', 'Scheme', 'Address', 'Comments']); r += 1
            _, PFw, *_ = self._s()
            warn_fill = PFw(start_color=P['warn'], end_color=P['warn'], fill_type='solid')
            for _, row in overlap.iterrows():
                rf = warn_fill
                self._dc(ws,r,1,row['job_id'],fill=rf,bold=True)
                self._dc(ws,r,2,row['gang_short'],fill=rf)
                self._dc(ws,r,3,row['abort_reason'],fill=rf)
                self._dc(ws,r,4,row['jv'],fill=rf,fmt=self.GF)
                self._dc(ws,r,5,row['wv'],fill=rf,fmt=self.GF)
                self._dc(ws,r,6,row['scheme'],fill=rf)
                self._dc(ws,r,7,str(row['address'])[:60] if row['address'] else '',fill=rf)
                self._dc(ws,r,8,str(row['comments'])[:80] if row['comments'] else '',fill=rf)
                r += 1
            r += 2

        # Section 4: ALL aborted jobs - full detail
        self._sec(ws, r, f"ALL ABORTED JOBS -- FULL DETAIL ({n_ab} jobs)"); r += 1
        all_hdrs = ['Job ID', 'Client Ref', 'Gang', 'Supervisor', 'Abort Reason', 'Job Status',
                    'Job Value (from Whereabouts)', 'In Whereabouts?', 'Scheme', 'Address', 'Comments', 'On Hold']
        self._wh(ws, r, all_hdrs); r += 1
        _, PFg, *_ = self._s()
        coa_fill = PFg(start_color=P['g'], end_color=P['g'], fill_type='solid')
        plan_fill = PFg(start_color=P['a'], end_color=P['a'], fill_type='solid')
        access_fill = PFg(start_color=P['r'], end_color=P['r'], fill_type='solid')
        for _, row in ab.sort_values(['abort_reason', 'gang']).iterrows():
            reason = row['abort_reason']
            if 'COA' in reason or 'Complete' in reason: rf = coa_fill
            elif 'Planning' in reason: rf = plan_fill
            elif 'Access' in reason or 'Refused' in reason: rf = access_fill
            else: rf = self._rfill(r)
            self._dc(ws,r,1,row['job_id'],fill=rf,bold=True)
            self._dc(ws,r,2,str(row['client_ref']) if row['client_ref'] else '',fill=rf)
            self._dc(ws,r,3,row['gang_short'],fill=rf)
            self._dc(ws,r,4,str(row['supervisor']) if row['supervisor'] else '',fill=rf)
            self._dc(ws,r,5,reason,fill=rf,bold=True)
            self._dc(ws,r,6,str(row['job_status']) if row['job_status'] else '',fill=rf)
            # Show value if in whereabouts, otherwise N/A
            if row['in_whereabouts'] and row['jv'] > 0:
                self._dc(ws,r,7,row['jv'],fill=rf,fmt=self.GF)
            else:
                self._dc(ws,r,7,'N/A - not in whereabouts',fill=rf)
            self._dc(ws,r,8,'Yes' if row['in_whereabouts'] else 'No',fill=rf)
            self._dc(ws,r,9,str(row['scheme']) if row['scheme'] else '',fill=rf)
            self._dc(ws,r,10,str(row['address'])[:60] if row['address'] else '',fill=rf)
            self._dc(ws,r,11,str(row['comments'])[:80] if row['comments'] else '',fill=rf)
            self._dc(ws,r,12,'Yes' if row['on_hold'] else 'No',fill=rf)
            r += 1

        ws.auto_filter.ref = f"A{r - len(ab)}:L{r-1}"
        ws.freeze_panes = 'A6'; self._aw(ws, mx=40)
        ws.column_dimensions['A'].width = 14; ws.column_dimensions['C'].width = 28
        ws.column_dimensions['E'].width = 28; ws.column_dimensions['J'].width = 35
        ws.column_dimensions['K'].width = 40

    # -- 9. All Jobs --
    def _all_jobs(self):
        from openpyxl.utils import get_column_letter
        F, PF, *_ = self._s()
        ws = self.wb.create_sheet("All Jobs (Sorted)"); ws.sheet_view.showGridLines = False
        keep = [c for c in ['Job ID','Client Ref 1','Contract Number','Work Type','Gang','Address','Postcode',
                'Scheduled From','Job Value','Whereabouts Value','Finish Status','Finish Comments'] if c in self.e.data.columns]
        dfs = self.e.data[keep + ['_shared','_orig_jv']].sort_values('Job Value', ascending=False).reset_index(drop=True)
        # Display columns: keep + Shared? + Original Value (only shown if shared jobs exist)
        has_shared = self.e.duplicates and len(self.e.duplicates) > 0
        disp_hdrs = keep + (['Shared Job?', 'Original Value'] if has_shared else [])
        ncols = len(disp_hdrs)
        lcl = get_column_letter(ncols)
        self._title(ws,f"ALL {len(dfs)} JOBS -- SORTED BY VALUE (HIGHEST FIRST)",lcl)
        self._sub(ws,"Filter and sort by any column" + (" | Shared jobs flagged - value assigned to first gang" if has_shared else ""),lcl)
        self._wh(ws,4,disp_hdrs)
        warn_fill = PF(start_color=P['warn'],end_color=P['warn'],fill_type='solid')
        for ri, (_, row) in enumerate(dfs.iterrows(), 5):
            is_shared = bool(row.get('_shared', False))
            rf = warn_fill if is_shared and row['Job Value'] == 0 else self._rfill(ri)
            for ci, col in enumerate(keep, 1):
                val = row[col]; cell = ws.cell(row=ri, column=ci)
                if pd.isna(val): cell.value = ''
                elif isinstance(val, pd.Timestamp): cell.value = val; cell.number_format = 'DD/MM/YYYY'
                elif isinstance(val, float): cell.value = round(val, 2)
                else: cell.value = val
                cell.font = F(name='Arial',size=10); cell.fill = rf
                from openpyxl.styles import Alignment as Al
                cell.alignment = Al(horizontal='center',vertical='center')
                if col in ('Job Value','Whereabouts Value'): cell.number_format = self.GF
            if has_shared:
                shared_ci = len(keep) + 1
                orig_ci = len(keep) + 2
                if is_shared:
                    if row['Job Value'] == 0:
                        self._dc(ws,ri,shared_ci,'ZEROED (shared)',fill=rf,bold=True)
                    else:
                        self._dc(ws,ri,shared_ci,'VALUE OWNER',fill=rf,bold=True)
                    self._dc(ws,ri,orig_ci,round(float(row['_orig_jv']),2),fill=rf,fmt=self.GF)
                else:
                    self._dc(ws,ri,shared_ci,'',fill=rf)
                    self._dc(ws,ri,orig_ci,'',fill=rf)
        last = 4 + len(dfs); ws.auto_filter.ref = f"A4:{lcl}{last}"; ws.freeze_panes = 'A5'
        self._aw(ws,mn=10,mx=30)
        gi = keep.index('Gang') + 1 if 'Gang' in keep else None
        if gi: ws.column_dimensions[get_column_letter(gi)].width = 36
        if has_shared: ws.column_dimensions[get_column_letter(len(keep)+1)].width = 18

    # -- 10. Raw Data --
    def _raw(self):
        from openpyxl.utils import get_column_letter
        F, *_ = self._s()
        ws = self.wb.create_sheet("Raw Data"); ws.sheet_view.showGridLines = True
        # Filter out internal columns
        raw = self.e.raw[[c for c in self.e.raw.columns if not c.startswith('_')]]
        for ci, cn in enumerate(raw.columns, 1):
            c = ws.cell(row=1,column=ci,value=cn); c.font = self._hft(); c.fill = self._hf()
            from openpyxl.styles import Alignment as Al
            c.alignment = Al(horizontal='center',vertical='center',wrap_text=True)
        for ri, (_, rd) in enumerate(raw.iterrows(), 2):
            for ci, val in enumerate(rd, 1):
                cell = ws.cell(row=ri,column=ci)
                if pd.isna(val): cell.value = ''
                elif isinstance(val, pd.Timestamp): cell.value = val; cell.number_format = 'DD/MM/YYYY HH:MM'
                elif isinstance(val, float): cell.value = round(val, 2)
                else: cell.value = val
                cell.font = F(name='Arial',size=10)
        lr = 1 + len(raw); lcl = get_column_letter(len(raw.columns))
        ws.auto_filter.ref = f"A1:{lcl}{lr}"; ws.freeze_panes = 'A2'; self._aw(ws,mn=8,mx=25)


# ===================================================================
# GUI
# ===================================================================

class App:
    def __init__(self):
        self.root = tk.Tk(); self.root.title("Whereabouts Values Analyzer v1.5")
        self.root.geometry("760x680"); self.root.resizable(True, True)
        try: self.root.configure(bg='#1a1a2e')
        except: pass
        self.engine = Engine(); self.fps = []; self.aborted_fps = []; self.last_out = None
        self.source_type = 'whereabouts'
        self._load_prefs(); self._ui()

    def _prefs_path(self): return os.path.join(os.path.expanduser('~'), '.whereabouts_prefs.txt')
    def _load_prefs(self):
        self._last_dir = None
        try:
            if os.path.exists(self._prefs_path()):
                for line in open(self._prefs_path()):
                    if line.startswith('last_dir='):
                        d = line.strip().split('=',1)[1]
                        if os.path.isdir(d): self._last_dir = d
        except: pass
    def _save_prefs(self):
        try:
            with open(self._prefs_path(),'w') as f:
                if self.fps: f.write(f"last_dir={os.path.dirname(self.fps[0])}\n")
        except: pass

    def _ui(self):
        sty = ttk.Style()
        try: sty.theme_use('clam')
        except: pass
        for n, c in [('T.TLabel',{'font':('Arial',16,'bold'),'foreground':'#F4B942','background':'#1a1a2e'}),
            ('S.TLabel',{'font':('Arial',10),'foreground':'#a0a0b0','background':'#1a1a2e'}),
            ('G.TLabel',{'font':('Arial',10),'foreground':'#50fa7b','background':'#1a1a2e'}),
            ('D.TFrame',{'background':'#1a1a2e'}),('C.TFrame',{'background':'#16213e'}),
            ('A.TButton',{'font':('Arial',11,'bold'),'padding':8})]: sty.configure(n,**c)
        m = ttk.Frame(self.root,style='D.TFrame',padding=20); m.pack(fill='both',expand=True)
        ttk.Label(m,text="WHEREABOUTS VALUES ANALYZER v1.5",style='T.TLabel').pack(pady=(0,2))
        ttk.Label(m,text="League Table | Value Analysis | Missing Whereabouts | PM Breakdown | Share",style='S.TLabel').pack(pady=(0,14))
        ff = ttk.Frame(m,style='C.TFrame',padding=12); ff.pack(fill='x',pady=(0,10))
        # Source type selector
        stf = ttk.Frame(ff,style='C.TFrame'); stf.pack(fill='x',pady=(0,8))
        ttk.Label(stf,text="Source Type:",font=('Arial',10,'bold'),foreground='#a0a0b0',background='#16213e').pack(side='left',padx=(0,8))
        self.src_var = tk.StringVar(value='whereabouts')
        ttk.Radiobutton(stf,text="Whereabouts Export",variable=self.src_var,value='whereabouts',command=self._src_changed).pack(side='left',padx=(0,12))
        ttk.Radiobutton(stf,text="Planned Report",variable=self.src_var,value='planned',command=self._src_changed).pack(side='left')
        ttk.Label(ff,text="Source Files:",font=('Arial',10,'bold'),foreground='#a0a0b0',background='#16213e').pack(anchor='w')
        pf = ttk.Frame(ff,style='C.TFrame'); pf.pack(fill='x',pady=(4,0))
        self.fv = tk.StringVar(value="No files selected")
        ttk.Label(pf,textvariable=self.fv,font=('Arial',9),foreground='#50fa7b',background='#16213e',wraplength=560).pack(side='left',fill='x',expand=True)
        ttk.Button(pf,text="Browse...",command=self._browse).pack(side='right')
        # Aborted jobs file
        ttk.Label(ff,text="Aborted Jobs (optional):",font=('Arial',10,'bold'),foreground='#a0a0b0',background='#16213e').pack(anchor='w',pady=(8,0))
        af = ttk.Frame(ff,style='C.TFrame'); af.pack(fill='x',pady=(4,0))
        self.afv = tk.StringVar(value="No aborted file selected")
        ttk.Label(af,textvariable=self.afv,font=('Arial',9),foreground='#F4B942',background='#16213e',wraplength=560).pack(side='left',fill='x',expand=True)
        ttk.Button(af,text="Browse...",command=self._browse_aborted).pack(side='right')
        of = ttk.Frame(m,style='C.TFrame',padding=12); of.pack(fill='x',pady=(0,10))
        self.desk = tk.BooleanVar(value=True); self.opn = tk.BooleanVar(value=True)
        cbf = ttk.Frame(of,style='C.TFrame'); cbf.pack(fill='x')
        ttk.Checkbutton(cbf,text="Save to Desktop",variable=self.desk).pack(side='left',padx=(0,16))
        ttk.Checkbutton(cbf,text="Open after",variable=self.opn).pack(side='left')
        self.btn = ttk.Button(m,text="ANALYSE & EXPORT",style='A.TButton',command=self._run); self.btn.pack(pady=(4,10))
        self.pb = ttk.Progressbar(m,mode='determinate',length=400); self.pb.pack(fill='x',pady=(0,6))
        self.sv = tk.StringVar(value="Ready - select Whereabouts Excel export(s)")
        ttk.Label(m,textvariable=self.sv,style='G.TLabel',wraplength=700).pack(anchor='w')
        self.sf = ttk.Frame(m,style='C.TFrame',padding=8); self.sf.pack(fill='x',pady=(6,0))
        ttk.Label(self.sf,text="Share:",font=('Arial',10,'bold'),foreground='#a0a0b0',background='#16213e').pack(side='left',padx=(0,8))
        ttk.Button(self.sf,text="Clipboard",command=self._clip).pack(side='left',padx=(0,6))
        ttk.Button(self.sf,text="Outlook",command=self._outlook).pack(side='left',padx=(0,6))
        ttk.Button(self.sf,text="Teams",command=self._teams).pack(side='left',padx=(0,6))
        self.sf.pack_forget()
        lf = ttk.Frame(m,style='C.TFrame',padding=4); lf.pack(fill='both',expand=True,pady=(6,0))
        self.log = tk.Text(lf,height=12,bg='#0f3460',fg='#e0e0e0',font=('Consolas',9),relief='flat',wrap='word')
        self.log.pack(fill='both',expand=True)

    def _l(self, msg):
        ts = datetime.now().strftime('%H:%M:%S')
        self.log.insert('end',f"[{ts}] {msg}\n"); self.log.see('end'); self.root.update_idletasks()
    def _p(self, v, text=None):
        self.pb['value'] = v
        if text: self.sv.set(text)
        self.root.update_idletasks()
    def _src_changed(self):
        st = self.src_var.get()
        self.source_type = st
        self.fps = []; self.fv.set("No files selected")
        if st == 'planned':
            self.sv.set("Ready - select Planning Board Excel (with Planned Report tab)")
        else:
            self.sv.set("Ready - select Whereabouts Excel export(s)")
    def _browse(self):
        init = self._last_dir or (str(Path.home()/'Downloads') if (Path.home()/'Downloads').exists() else str(Path.home()))
        paths = filedialog.askopenfilenames(title="Select Whereabouts Export(s)",initialdir=init,filetypes=[('Excel','*.xlsx *.xls'),('All','*.*')])
        if paths:
            self.fps = list(paths)
            self.fv.set(f"{len(self.fps)} file{'s' if len(self.fps)>1 else ''}: {', '.join(os.path.basename(p) for p in self.fps)}")
    def _browse_aborted(self):
        init = self._last_dir or (str(Path.home()/'Downloads') if (Path.home()/'Downloads').exists() else str(Path.home()))
        paths = filedialog.askopenfilenames(title="Select Aborted Jobs Export(s)",initialdir=init,filetypes=[('Excel','*.xlsx *.xls'),('All','*.*')])
        if paths:
            self.aborted_fps = list(paths)
            self.afv.set(f"{len(self.aborted_fps)} file{'s' if len(self.aborted_fps)>1 else ''}: {', '.join(os.path.basename(p) for p in self.aborted_fps)}")
    def _clip(self):
        try: self.root.clipboard_clear(); self.root.clipboard_append(self.engine.share_text()); self.root.update(); self._l("Copied to clipboard")
        except Exception as ex: self._l(f"Error: {ex}")

    def _outlook(self):
        try:
            s = self.engine.summary
            subj = f"{s['contract']} {s['region']} - Whereabouts Values {datetime.now().strftime('%d/%m/%Y')}"
            if sys.platform == 'win32':
                html = self.engine.share_html()
                html_path = os.path.join(os.path.expanduser('~'), '_whereabouts_email.html')
                # Write as ASCII - our HTML uses only &pound; entities, no unicode
                with open(html_path, 'w', encoding='ascii') as f: f.write(html)
                al = f'    .Attachments.Add "{self.last_out}"' if self.last_out and os.path.exists(self.last_out) else ''
                # OpenTextFile flag: 0 = ASCII (TristateFalse), NOT -1 which is UTF-16
                vbs = ('On Error Resume Next\nDim o, m, fso, ts, htmlBody\n'
                       'Set fso = CreateObject("Scripting.FileSystemObject")\n'
                       f'Set ts = fso.OpenTextFile("{html_path}", 1, False, 0)\nhtmlBody = ts.ReadAll\nts.Close\n'
                       'Set o = CreateObject("Outlook.Application")\nSet m = o.CreateItem(0)\nWith m\n'
                       f'    .Subject = "{subj}"\n    .HTMLBody = htmlBody\n{al}\n    .Display\nEnd With\n'
                       'If Err.Number <> 0 Then WScript.Quit 1\n')
                vp = os.path.join(os.path.expanduser('~'), '_whereabouts_email.vbs')
                with open(vp, 'w', encoding='ascii', errors='replace') as f: f.write(vbs)
                try:
                    import subprocess; result = subprocess.run(['wscript', vp], shell=True, timeout=10)
                    if result.returncode == 0: self._l("Outlook opened with report attached")
                    else: raise Exception("VBS error")
                except:
                    self._l("VBS failed, trying PowerShell...")
                    att = f'$m.Attachments.Add("{self.last_out}")' if self.last_out and os.path.exists(self.last_out) else ''
                    ps = (f'$html = Get-Content -Path "{html_path}" -Raw -Encoding ASCII\n'
                          '$o = New-Object -ComObject Outlook.Application\n$m = $o.CreateItem(0)\n'
                          f'$m.Subject = "{subj}"\n$m.HTMLBody = $html\n{att}\n$m.Display()\n')
                    ps_path = os.path.join(os.path.expanduser('~'), '_whereabouts_email.ps1')
                    with open(ps_path, 'w', encoding='utf-8') as f: f.write(ps)
                    try:
                        import subprocess; subprocess.Popen(['powershell','-ExecutionPolicy','Bypass','-File',ps_path])
                        self._l("Outlook opened via PowerShell")
                    except:
                        import urllib.parse; os.startfile(f'mailto:?subject={urllib.parse.quote(subj)}&body={urllib.parse.quote(self.engine.share_text())}')
                        self._l("Mailto opened - attach report manually")
            else:
                self.root.clipboard_clear(); self.root.clipboard_append(self.engine.share_text()); self.root.update()
                self._l(f"Copied - paste into email. Attach: {self.last_out}")
        except Exception as ex: self._l(f"Error: {ex}")

    def _teams(self):
        cfg = os.path.join(os.path.expanduser('~'), '.whereabouts_webhook.txt'); saved = ''
        try:
            if os.path.exists(cfg):
                with open(cfg) as f: saved = f.read().strip()
        except: pass
        dlg = tk.Toplevel(self.root); dlg.title("Teams Webhook"); dlg.geometry("520x200")
        try: dlg.configure(bg='#1a1a2e')
        except: pass
        ttk.Label(dlg,text="Webhook URL:",font=('Arial',10,'bold'),foreground='#a0a0b0',background='#1a1a2e').pack(pady=(16,4),padx=16,anchor='w')
        uv = tk.StringVar(value=saved); ttk.Entry(dlg,textvariable=uv,font=('Arial',10),width=60).pack(padx=16,fill='x')
        sv = tk.BooleanVar(value=True); ttk.Checkbutton(dlg,text="Remember",variable=sv).pack(pady=(8,4),padx=16,anchor='w')
        sl = ttk.Label(dlg,text="",foreground='#50fa7b',background='#1a1a2e'); sl.pack(pady=4)
        def send():
            url = uv.get().strip()
            if not url: sl.configure(text="Enter URL",foreground='#e94560'); return
            try:
                import urllib.request, json
                payload = json.dumps({"type":"message","attachments":[{"contentType":"application/vnd.microsoft.card.adaptive",
                    "content":{"$schema":"http://adaptivecards.io/schemas/adaptive-card.json","type":"AdaptiveCard","version":"1.4",
                        "body":[{"type":"TextBlock","weight":"bolder","size":"large","color":"accent",
                                 "text":f"{self.engine.summary['contract']} {self.engine.summary['region']} - Whereabouts Values"},
                                {"type":"TextBlock","text":self.engine.share_text(),"wrap":True,"fontType":"monospace","size":"small"}]}}]})
                req = urllib.request.Request(url,data=payload.encode('utf-8'),headers={'Content-Type':'application/json'})
                resp = urllib.request.urlopen(req,timeout=10)
                if sv.get():
                    with open(cfg,'w') as f: f.write(url)
                sl.configure(text=f"Sent! ({resp.getcode()})",foreground='#50fa7b'); self._l("Posted to Teams"); dlg.after(1500,dlg.destroy)
            except Exception as ex: sl.configure(text=str(ex)[:60],foreground='#e94560')
        ttk.Button(dlg,text="Send",command=send).pack(pady=(4,16))

    def _run(self):
        if not self.fps: messagebox.showerror("Error","Select file(s)"); return
        self.engine = Engine()  # Reset engine for fresh run
        self.btn.configure(state='disabled'); self.log.delete('1.0','end'); self.sf.pack_forget()
        try:
            is_planned = self.src_var.get() == 'planned'
            self._p(5,"Loading files..."); self._l(f"Loading {len(self.fps)} file(s) ({'Planned Report' if is_planned else 'Whereabouts'} mode)...")
            if is_planned:
                df = self.engine.load_planned_report(self.fps)
                self._l(f"Loaded {len(df)} rows from Planned Report")
            else:
                df = self.engine.load_files(self.fps)
                self._l(f"Loaded {len(df)} rows")
            self._p(15,"Analysing..."); self.engine.analyse(df); s = self.engine.summary
            # Planned report extra log output
            if is_planned and self.engine.planned_stats:
                ps = self.engine.planned_stats
                self._l(f"\nPLANNED REPORT SUMMARY:")
                self._l(f"  Planned Value: {GBP}{ps['total_planned']:,.2f}")
                self._l(f"  Measured Value: {GBP}{ps['total_measured']:,.2f}")
                self._l(f"  Remaining: {GBP}{ps['total_remaining']:,.2f}")
                self._l(f"  Measure Rate: {ps['measure_rate']:.1f}%")
                self._l(f"  Site Cleared: {ps['n_site_cleared']} | Started: {ps['n_started']} | Measures Logged: {ps['n_measures_logged']}")
                self._l(f"  Missing Measures: {ps['n_missing_measures']} | Aborted: {ps['n_aborted']} | CVs: {ps['n_cv']} ({GBP}{ps['total_cv_value']:,.2f})")
                self._l(f"  On Hold: {ps['n_on_hold']} | Unscheduled: {ps['n_unscheduled']} | Ready To Invoice: {ps['n_ready_invoice']}")
            # Load aborted jobs if provided
            if self.aborted_fps:
                self._p(20,"Loading aborted jobs...")
                self._l(f"Loading {len(self.aborted_fps)} aborted file(s)...")
                self.engine.load_aborted(self.aborted_fps)
                self.engine.analyse_aborted()
                if self.engine.aborted_analysis is not None:
                    n_ab = len(self.engine.aborted_analysis)
                    n_ov = int(self.engine.aborted_analysis['in_whereabouts'].sum())
                    jv_ov = round(float(self.engine.aborted_analysis[self.engine.aborted_analysis['in_whereabouts']]['jv'].sum()), 2)
                    self._l(f"Aborted: {n_ab} jobs | {n_ov} also in whereabouts ({GBP}{jv_ov:,.2f} value)")
                    for reason, rs in sorted(self.engine.aborted_by_reason.items(), key=lambda x: -x[1]['count']):
                        self._l(f"   {reason}: {rs['count']} jobs")
            if self.engine.duplicates:
                self._l(f"\n!! {len(self.engine.duplicates)} SHARED JOBS (value assigned to first gang, others zeroed):")
                for d in self.engine.duplicates:
                    others = [g for g in d['gang_shorts'] if g != d['owner']]
                    self._l(f"   Job {d['job_id']}: {GBP}{d['orig_val']:,.2f} -> {d['owner']} | {', '.join(others)} get {GBP}0.00 | {d['scheme']} | {d['day']}")
            if s['missing_wv_count'] > 0:
                mp = round(s['missing_wv_value']/s['total_jv']*100) if s['total_jv']>0 else 0
                self._l(f"\n!! MISSING WHEREABOUTS: {s['missing_wv_count']} jobs = {GBP}{s['missing_wv_value']:,.2f} ({mp}% of total)")
                for gang, mg in sorted(self.engine.missing_wv_by_gang.items(), key=lambda x: -x[1]['missing_jv']):
                    self._l(f"   {mg['short']}: {mg['missing_count']}/{mg['total_count']} ({mg['missing_pct']:.0f}%) = {GBP}{mg['missing_jv']:,.2f}")
            self._p(25,""); self._l(f"\n{s['contract']} | {s['region']} | {s['n_days']} days | {s['n_pms']} PMs | {s['n_gangs']} gangs")
            self._l(f"Total: {s['total_jobs']} jobs | {GBP}{s['total_jv']:,.2f} value | {GBP}{s['total_wv']:,.2f} whereabouts | {GBP}{s['avg_jv']:,.2f} avg")
            self._l("")
            for pm, pst in sorted(self.engine.pm_stats.items(), key=lambda x: x[1]['jv'], reverse=True):
                self._l(f"  {pm}: {pst['n']} jobs | {GBP}{pst['jv']:,.2f} | {GBP}{pst['avg']:,.2f} avg")
            self._l("\nGANG LEAGUE TABLE:")
            for i, (gang, st) in enumerate(sorted(self.engine.gang_stats.items(), key=lambda x: x[1]['jv'], reverse=True), 1):
                self._l(f"  #{i} {st['short']} -- {st['n']} jobs -- {GBP}{st['jv']:,.2f} -- {GBP}{st['avg']:,.2f} avg")
            def on_prog(name,step,total): self._p(30+int(step/total*60),f"Sheet {step}/{total}: {name}...")
            prefix = "planned_report" if is_planned else "whereabouts_values"
            fn = f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            out = str((Path.home()/'Desktop'/fn) if self.desk.get() and (Path.home()/'Desktop').exists() else Path(self.fps[0]).parent/fn)
            Report(self.engine).build(out,progress_cb=on_prog); self.last_out = out; self._save_prefs()
            try: sz = os.path.getsize(out)/1024; self._l(f"\nSaved: {out} ({sz:.0f} KB)" if sz < 1024 else f"\nSaved: {out} ({sz/1024:.1f} MB)")
            except: self._l(f"\nSaved: {out}")
            self._p(100,f"Done - {os.path.basename(out)}"); self.sf.pack(fill='x',pady=(6,0))
            if self.opn.get():
                try:
                    if sys.platform=='win32': os.startfile(out)
                    elif sys.platform=='darwin': os.system(f'open "{out}"')
                    else: os.system(f'xdg-open "{out}"')
                except: pass
        except Exception as ex:
            self._p(0,f"Error: {ex}"); self._l(f"\nERROR: {ex}\n{traceback.format_exc()}"); messagebox.showerror("Failed",str(ex))
        finally: self.btn.configure(state='normal')
    def run(self): self.root.mainloop()


def run_cli(paths):
    print("WHEREABOUTS VALUES ANALYZER v1.5 -- CLI"); print("="*50)
    e = Engine(); df = e.load_files(paths); print(f"{len(df)} rows"); e.analyse(df); s = e.summary
    if e.duplicates:
        print(f"\n!! {len(e.duplicates)} SHARED JOBS (value assigned to first gang):")
        for d in e.duplicates:
            others = [g for g in d['gang_shorts'] if g != d['owner']]
            print(f"   Job {d['job_id']}: {GBP}{d['orig_val']:,.2f} -> {d['owner']} | {', '.join(others)} get {GBP}0.00")
    if s['missing_wv_count']>0:
        print(f"\n!! MISSING WHEREABOUTS: {s['missing_wv_count']} jobs = {GBP}{s['missing_wv_value']:,.2f}")
        for g, mg in sorted(e.missing_wv_by_gang.items(), key=lambda x: -x[1]['missing_jv']):
            print(f"   {mg['short']}: {mg['missing_count']}/{mg['total_count']} ({mg['missing_pct']:.0f}%) = {GBP}{mg['missing_jv']:,.2f}")
    print(f"\n{e.share_text()}")
    fn = f"whereabouts_values_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    for folder in [Path.home()/'Desktop', Path.home()/'Downloads', Path.cwd()]:
        if folder.exists():
            try: Report(e).build(str(folder/fn)); print(f"\nSaved: {folder/fn}"); return
            except PermissionError: continue

def main():
    if len(sys.argv) > 1:
        paths = [p for p in sys.argv[1:] if os.path.exists(p)]
        if not paths: print("No valid files"); sys.exit(1)
        run_cli(paths)
    elif HAS_TK:
        try: App().run()
        except Exception as ex: print(f"GUI failed: {ex}"); traceback.print_exc()
    else: print("Usage: python WHEREABOUTS_VALUES_v14.py <file1.xlsx> [file2.xlsx ...]")

if __name__ == '__main__': main()
